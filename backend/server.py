"""Spartans Growth League — FastAPI backend.
Handles auth (JWT + Emergent Google), gamification, CRM, challenges, leaderboard, admin.
"""
from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import os
import uuid
import math
import logging
import secrets
import httpx
from datetime import datetime, timezone, timedelta, date
from typing import Optional, List, Literal

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, UploadFile, File
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, EmailStr, Field, ConfigDict

from auth_utils import (
    hash_password, verify_password,
    create_access_token, create_refresh_token,
    set_auth_cookies, clear_auth_cookies,
    get_current_user, require_role,
    get_jwt_secret,
)
import jwt

# ---------- Setup ----------
mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

app = FastAPI(title="Spartans Growth League API")
api = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("spartans")

EMERGENT_AUTH_URL = "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data"

# ---------- Gamification helpers ----------
def level_from_xp(xp: int) -> int:
    if xp <= 0:
        return 1
    return int(math.floor(math.sqrt(xp / 100.0))) + 1


def xp_to_level(level: int) -> int:
    return (level - 1) ** 2 * 100


def xp_progress(xp: int):
    lvl = level_from_xp(xp)
    curr = xp_to_level(lvl)
    nxt = xp_to_level(lvl + 1)
    return {"level": lvl, "xp": xp, "current_level_xp": curr, "next_level_xp": nxt,
            "progress_pct": round(((xp - curr) / max(1, nxt - curr)) * 100, 1)}


XP_RULES = {
    "daily_checkin": 10,
    "prospect_added": 5,
    "prospect_won": 50,
    "followup_done": 8,
    "attendance": 15,
    "challenge_completed": 100,
    "mission_logged": 10,
    "mission_converted": 40,
}

# Master badge catalog (seeded)
BADGE_CATALOG = [
    {"key": "first_step", "name": "First Step", "description": "Complete your first daily check-in.", "icon": "Sword", "req_type": "checkins", "req_value": 1, "xp_reward": 25, "tier": "bronze"},
    {"key": "streak_7", "name": "Warrior Week", "description": "Maintain a 7-day check-in streak.", "icon": "Fire", "req_type": "streak", "req_value": 7, "xp_reward": 75, "tier": "silver"},
    {"key": "streak_30", "name": "Spartan Iron", "description": "Maintain a 30-day check-in streak.", "icon": "Flame", "req_type": "streak", "req_value": 30, "xp_reward": 300, "tier": "gold"},
    {"key": "prospect_10", "name": "Hunter", "description": "Add 10 prospects.", "icon": "Target", "req_type": "prospects", "req_value": 10, "xp_reward": 50, "tier": "bronze"},
    {"key": "prospect_50", "name": "Elite Recruiter", "description": "Add 50 prospects.", "icon": "Crosshair", "req_type": "prospects", "req_value": 50, "xp_reward": 250, "tier": "silver"},
    {"key": "closer_5", "name": "Closer", "description": "Convert 5 prospects to won.", "icon": "Trophy", "req_type": "prospects_won", "req_value": 5, "xp_reward": 200, "tier": "gold"},
    {"key": "followup_25", "name": "Persistent", "description": "Complete 25 follow-ups.", "icon": "Phone", "req_type": "followups_done", "req_value": 25, "xp_reward": 100, "tier": "silver"},
    {"key": "attendance_10", "name": "Committed", "description": "Attend 10 events.", "icon": "Calendar", "req_type": "attendance", "req_value": 10, "xp_reward": 120, "tier": "silver"},
    {"key": "level_5", "name": "Rising Spartan", "description": "Reach level 5.", "icon": "Star", "req_type": "level", "req_value": 5, "xp_reward": 150, "tier": "silver"},
    {"key": "level_10", "name": "Battle Master", "description": "Reach level 10.", "icon": "ShieldStar", "req_type": "level", "req_value": 10, "xp_reward": 500, "tier": "gold"},
]


# ---------- Pydantic Models ----------
class UserPublic(BaseModel):
    model_config = ConfigDict(extra="ignore")
    user_id: str
    email: EmailStr
    name: str
    role: str = "member"
    avatar_url: Optional[str] = None
    picture: Optional[str] = None
    xp: int = 0
    level: int = 1
    streak_current: int = 0
    streak_longest: int = 0
    last_checkin_date: Optional[str] = None
    team: Optional[str] = None
    badges: List[str] = []
    created_at: Optional[str] = None


import re
from pydantic import field_validator


def _clean_phone_10(v):
    if v is None or v == "":
        return None
    s = re.sub(r"\D", "", str(v))
    if len(s) != 10:
        raise ValueError("Mobile number must be exactly 10 digits")
    return s


def _reject_future_date(v):
    if v is None or v == "":
        return None
    try:
        d = date.fromisoformat(v)
    except Exception:
        raise ValueError("Invalid date (expected YYYY-MM-DD)")
    if d > date.today():
        raise ValueError("Date cannot be in the future")
    return v


# RegisterIn model with validators
class RegisterIn(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    name: str = Field(min_length=1)
    nexus_id: str = Field(min_length=3, max_length=40)
    phone: Optional[str] = None
    dob: Optional[str] = None
    gender: Optional[Literal["male", "female", "other", "prefer_not"]] = None
    city: Optional[str] = None
    state: Optional[str] = None

    @field_validator("phone")
    @classmethod
    def v_phone(cls, v):
        return _clean_phone_10(v)

    @field_validator("dob")
    @classmethod
    def v_dob(cls, v):
        return _reject_future_date(v)

    @field_validator("nexus_id")
    @classmethod
    def v_nexus_id(cls, v):
        s = str(v).strip()
        if len(s) < 3:
            raise ValueError("Nexus ID (Business Centre) must be at least 3 characters")
        return s


class LoginIn(BaseModel):
    email: EmailStr
    password: str


class GoogleSessionIn(BaseModel):
    session_id: str


class ProspectIn(BaseModel):
    name: str
    contact: Optional[str] = None
    status: Literal["new", "contacted", "qualified", "won", "lost"] = "new"
    notes: Optional[str] = None
    source: Optional[str] = None


class ProspectUpdate(BaseModel):
    name: Optional[str] = None
    contact: Optional[str] = None
    status: Optional[Literal["new", "contacted", "qualified", "won", "lost"]] = None
    notes: Optional[str] = None
    source: Optional[str] = None


FOLLOWUP_TIME_SLOTS = ["8-10am", "10-12pm", "2-4pm", "6-8pm"]


class FollowUpIn(BaseModel):
    title: str
    due_date: str  # ISO date
    time_slot: Optional[Literal["8-10am", "10-12pm", "2-4pm", "6-8pm"]] = None
    prospect_id: Optional[str] = None
    notes: Optional[str] = None


class FollowUpUpdate(BaseModel):
    status: Optional[Literal["pending", "done"]] = None
    title: Optional[str] = None
    due_date: Optional[str] = None
    time_slot: Optional[Literal["8-10am", "10-12pm", "2-4pm", "6-8pm"]] = None
    notes: Optional[str] = None


class AttendanceIn(BaseModel):
    event_name: str
    event_date: str  # ISO date
    event_type: Literal["meeting", "training", "webinar", "call"] = "meeting"
    notes: Optional[str] = None


class MissionIn(BaseModel):
    prospect_name: str = Field(min_length=1)
    mobile_number: Optional[str] = None
    notes: Optional[str] = None
    status: Literal["new", "followup", "converted"] = "new"
    lat: Optional[float] = None
    lng: Optional[float] = None
    photo_data: Optional[str] = None  # base64 data URL (data:image/jpeg;base64,...)
    accuracy: Optional[float] = None


class MissionUpdate(BaseModel):
    status: Optional[Literal["new", "followup", "converted"]] = None
    notes: Optional[str] = None


# --- Weekly Attendance / Seasons / Tasks models ---
class WeeklyEventIn(BaseModel):
    name: str
    clubs: List[Literal["converter", "believer", "builder", "decider"]] = []
    weekday: int = Field(ge=0, le=6)
    repeat_type: Literal["weekly", "once"] = "weekly"
    open_time: str = "08:00"
    lock_time: str = "22:00"
    active: bool = True
    completed: bool = False

class WeeklyEventUpdate(BaseModel):
    name: Optional[str] = None
    clubs: Optional[List[Literal["converter", "believer", "builder", "decider"]]] = None
    weekday: Optional[int] = Field(default=None, ge=0, le=6)
    repeat_type: Optional[Literal["weekly", "once"]] = None
    open_time: Optional[str] = None
    lock_time: Optional[str] = None
    active: Optional[bool] = None
    completed: Optional[bool] = None

class EventAttendanceMark(BaseModel):
    event_id: str
    event_date: str  # YYYY-MM-DD
    status: Literal["present", "absent", "na"]
    season_id: Optional[str] = None


class SeasonIn(BaseModel):
    name: str
    start_date: str  # YYYY-MM-DD
    end_date: str
    is_believer: bool = False
    total_pv: Optional[float] = Field(default=None, ge=0)
    total_earnings: Optional[float] = Field(default=None, ge=0)


class SeasonUpdate(BaseModel):
    name: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    is_believer: Optional[bool] = None
    total_pv: Optional[float] = Field(default=None, ge=0)
    total_earnings: Optional[float] = Field(default=None, ge=0)


class UserBusinessUpdate(BaseModel):
    """Super Admin can set per-user PV/Earnings for a season."""
    season_id: str
    pv: Optional[float] = Field(default=None, ge=0)
    earnings: Optional[float] = Field(default=None, ge=0)


class TaskIn(BaseModel):
    title: str
    description: str
    assigned_to: str  # user_id
    due_date: str  # YYYY-MM-DD
    xp_reward: int = Field(default=25, ge=0, le=1000)


class BelieverUpdate(BaseModel):
    is_believer: bool


class RewardIn(BaseModel):
    name: str
    description: Optional[str] = None
    cost_xp: int = Field(ge=1, le=100000)
    category: Literal["dinner", "movie", "outing", "voucher", "other"] = "other"
    stock: Optional[int] = None  # None = unlimited
    image_url: Optional[str] = None
    active: bool = True


class RewardUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    cost_xp: Optional[int] = Field(default=None, ge=1, le=100000)
    category: Optional[Literal["dinner", "movie", "outing", "voucher", "other"]] = None
    stock: Optional[int] = None
    image_url: Optional[str] = None
    active: Optional[bool] = None


class ProfileUpdate(BaseModel):
    # Basic
    name: Optional[str] = None
    phone: Optional[str] = None
    dob: Optional[str] = None  # YYYY-MM-DD
    gender: Optional[Literal["male", "female", "other", "prefer_not"]] = None
    marital_status: Optional[Literal["married", "unmarried"]] = None
    anniversary_date: Optional[str] = None  # YYYY-MM-DD
    anniversary_photo: Optional[str] = None  # base64 data URL
    avatar_url: Optional[str] = None
    avatar_photo: Optional[str] = None  # base64 alternative
    nexus_id: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    # Business
    joining_date: Optional[str] = None
    club_type: Optional[Literal["decider", "believer", "converter", "builder"]] = None
    position: Optional[str] = None
    # Personal
    favourite_food: Optional[str] = None
    favourite_place: Optional[str] = None
    favourite_hobby: Optional[str] = None
    bio: Optional[str] = None

    @field_validator("phone")
    @classmethod
    def v_phone(cls, v):
        return _clean_phone_10(v)

    @field_validator("dob", "joining_date")
    @classmethod
    def v_dates(cls, v):
        return _reject_future_date(v)


POSITION_BADGES = ["team_leader", "star_performer", "rising_star", "consistent_achiever", "top_recruiter"]


class PositionBadgesUpdate(BaseModel):
    badges: List[str]

    @field_validator("badges")
    @classmethod
    def v_badges(cls, v):
        invalid = [b for b in v if b not in POSITION_BADGES]
        if invalid:
            raise ValueError(f"Invalid badges: {invalid}")
        return v


class GoalIn(BaseModel):
    title: str = Field(min_length=1)
    target: int = Field(ge=1)
    period: Literal["weekly", "monthly"] = "weekly"
    xp_reward: int = Field(default=50, ge=0, le=1000)


class GoalProgress(BaseModel):
    progress: int = Field(ge=0)


class AddMemberIn(BaseModel):
    email: EmailStr
    name: str = Field(min_length=1)
    password: str = Field(min_length=6)
    phone: Optional[str] = None

    @field_validator("phone")
    @classmethod
    def v_phone(cls, v):
        return _clean_phone_10(v)


class ChallengeIn(BaseModel):
    title: str
    description: str
    type: Literal["weekly", "monthly"] = "weekly"
    goal_type: Literal["checkins", "prospects", "followups", "attendance", "xp"]
    goal: int
    start_date: str
    end_date: str
    xp_reward: int = 100
    badge_reward: Optional[str] = None


class RoleUpdate(BaseModel):
    role: Literal["super_admin", "team_leader", "member"]


class TeamIn(BaseModel):
    name: str
    leader_id: Optional[str] = None


class TeamUpdate(BaseModel):
    name: Optional[str] = None
    leader_id: Optional[str] = None


class TeamAssign(BaseModel):
    user_id: str
    is_leader: bool = False


# ---------- Utility ----------
def _iso(dt: datetime) -> str:
    return dt.astimezone(timezone.utc).isoformat()


async def award_xp(user_id: str, amount: int, reason: str):
    """Add XP, update level, check badges. Returns event summary."""
    user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not user:
        return None
    new_xp = int(user.get("xp", 0)) + int(amount)
    new_level = level_from_xp(new_xp)
    old_level = int(user.get("level", 1))
    await db.users.update_one({"user_id": user_id}, {"$set": {"xp": new_xp, "level": new_level}})
    await db.xp_events.insert_one({
        "event_id": str(uuid.uuid4()),
        "user_id": user_id,
        "amount": amount,
        "reason": reason,
        "created_at": _iso(datetime.now(timezone.utc)),
    })
    unlocked = await check_and_unlock_badges(user_id)
    leveled_up = new_level > old_level
    return {"xp": new_xp, "level": new_level, "leveled_up": leveled_up, "unlocked_badges": unlocked}


async def check_and_unlock_badges(user_id: str) -> list:
    user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not user:
        return []
    owned = set(user.get("badges", []))
    unlocked = []
    checkins = await db.checkins.count_documents({"user_id": user_id})
    prospects = await db.prospects.count_documents({"user_id": user_id})
    prospects_won = await db.prospects.count_documents({"user_id": user_id, "status": "won"})
    followups_done = await db.followups.count_documents({"user_id": user_id, "status": "done"})
    attendance = await db.attendance.count_documents({"user_id": user_id})
    for b in BADGE_CATALOG:
        if b["key"] in owned:
            continue
        cond = False
        rt = b["req_type"]
        if rt == "checkins":
            cond = checkins >= b["req_value"]
        elif rt == "streak":
            cond = user.get("streak_current", 0) >= b["req_value"]
        elif rt == "prospects":
            cond = prospects >= b["req_value"]
        elif rt == "prospects_won":
            cond = prospects_won >= b["req_value"]
        elif rt == "followups_done":
            cond = followups_done >= b["req_value"]
        elif rt == "attendance":
            cond = attendance >= b["req_value"]
        elif rt == "level":
            cond = user.get("level", 1) >= b["req_value"]
        if cond:
            await db.users.update_one({"user_id": user_id}, {"$addToSet": {"badges": b["key"]}, "$inc": {"xp": b["xp_reward"]}})
            await db.user_badges.insert_one({"user_id": user_id, "badge_key": b["key"], "unlocked_at": _iso(datetime.now(timezone.utc))})
            unlocked.append(b)
            owned.add(b["key"])
    if unlocked:
        u2 = await db.users.find_one({"user_id": user_id}, {"_id": 0})
        await db.users.update_one({"user_id": user_id}, {"$set": {"level": level_from_xp(u2.get("xp", 0))}})
    return unlocked


# ---------- Auth Endpoints ----------
async def _login_lockout_check(identifier: str):
    doc = await db.login_attempts.find_one({"identifier": identifier})
    if doc and doc.get("locked_until"):
        lu = doc["locked_until"]
        if isinstance(lu, str):
            lu = datetime.fromisoformat(lu)
        if lu.tzinfo is None:
            lu = lu.replace(tzinfo=timezone.utc)
        if lu > datetime.now(timezone.utc):
            raise HTTPException(status_code=429, detail="Too many failed attempts. Try again later.")


async def _login_record_failure(identifier: str):
    doc = await db.login_attempts.find_one({"identifier": identifier}) or {"identifier": identifier, "attempts": 0}
    attempts = doc.get("attempts", 0) + 1
    update = {"attempts": attempts, "identifier": identifier}
    if attempts >= 5:
        update["locked_until"] = _iso(datetime.now(timezone.utc) + timedelta(minutes=15))
        update["attempts"] = 0
    await db.login_attempts.update_one({"identifier": identifier}, {"$set": update}, upsert=True)


async def _login_clear(identifier: str):
    await db.login_attempts.delete_one({"identifier": identifier})


def _user_to_public(u: dict) -> dict:
    if not u:
        return u
    u.pop("password_hash", None)
    u.pop("_id", None)
    return u


@api.post("/auth/register")
async def register(payload: RegisterIn, response: Response):
    email = payload.email.lower().strip()
    exists = await db.users.find_one({"email": email})
    if exists:
        raise HTTPException(status_code=400, detail="Email already registered")
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    doc = {
        "user_id": user_id,
        "email": email,
        "name": payload.name.strip(),
        "password_hash": hash_password(payload.password),
        "role": "member",
        "avatar_url": None,
        "picture": None,
        "nexus_id": payload.nexus_id.strip(),
        "xp": 0,
        "level": 1,
        "streak_current": 0,
        "streak_longest": 0,
        "last_checkin_date": None,
        "team": "Alpha",
        "phone": payload.phone,
        "dob": payload.dob,
        "gender": payload.gender,
        "city": payload.city,
        "state": payload.state,
        "badges": [],
        "created_at": _iso(datetime.now(timezone.utc)),
        "active": True,
    }
    await db.users.insert_one(doc)
    # Auto-assign active admin goal templates to the new member
    try:
        active_templates = await db.goal_templates.find({"active": True}, {"_id": 0}).to_list(200)
        for tpl in active_templates:
            await _sync_template_to_users(tpl)
    except Exception as e:
        logger.warning(f"Failed to auto-assign goal templates: {e}")
    at = create_access_token(user_id, email)
    rt = create_refresh_token(user_id)
    set_auth_cookies(response, at, rt)
    return {"user": _user_to_public(doc), "access_token": at}


@api.post("/auth/login")
async def login(payload: LoginIn, request: Request, response: Response):
    email = payload.email.lower().strip()
    xff = request.headers.get("x-forwarded-for", "")
    ip = xff.split(",")[0].strip() if xff else (request.client.host if request.client else "unknown")
    ident = f"{ip}:{email}"
    await _login_lockout_check(ident)
    user = await db.users.find_one({"email": email})
    if not user or not user.get("password_hash") or not verify_password(payload.password, user["password_hash"]):
        await _login_record_failure(ident)
        raise HTTPException(status_code=401, detail="Invalid email or password")
    await _login_clear(ident)
    at = create_access_token(user["user_id"], email)
    rt = create_refresh_token(user["user_id"])
    set_auth_cookies(response, at, rt)
    return {"user": _user_to_public(user), "access_token": at}


@api.post("/auth/logout")
async def logout(request: Request, response: Response):
    session_token = request.cookies.get("session_token")
    if session_token:
        await db.user_sessions.delete_one({"session_token": session_token})
    clear_auth_cookies(response)
    return {"ok": True}


@api.get("/auth/me")
async def me(request: Request):
    user = await get_current_user(request, db)
    return _user_to_public(user)


@api.post("/auth/refresh")
async def refresh(request: Request, response: Response):
    rt = request.cookies.get("refresh_token")
    if not rt:
        raise HTTPException(status_code=401, detail="No refresh token")
    try:
        payload = jwt.decode(rt, get_jwt_secret(), algorithms=["HS256"])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"user_id": payload["sub"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        at = create_access_token(user["user_id"], user["email"])
        new_rt = create_refresh_token(user["user_id"])
        set_auth_cookies(response, at, new_rt)
        return {"ok": True}
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid refresh token")


@api.post("/auth/google-session")
async def google_session(payload: GoogleSessionIn, response: Response):
    """Exchange Emergent session_id for a stored session_token cookie."""
    async with httpx.AsyncClient(timeout=15.0) as http:
        r = await http.get(EMERGENT_AUTH_URL, headers={"X-Session-ID": payload.session_id})
        if r.status_code != 200:
            raise HTTPException(status_code=401, detail="Invalid Google session")
        data = r.json()

    email = (data.get("email") or "").lower()
    name = data.get("name") or "Spartan"
    picture = data.get("picture")
    session_token = data.get("session_token")
    if not email or not session_token:
        raise HTTPException(status_code=401, detail="Incomplete Google session data")

    user = await db.users.find_one({"email": email})
    if not user:
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        user = {
            "user_id": user_id, "email": email, "name": name, "password_hash": None,
            "role": "member", "picture": picture, "avatar_url": picture,
            "xp": 0, "level": 1, "streak_current": 0, "streak_longest": 0,
            "last_checkin_date": None, "team": "Alpha", "badges": [],
            "created_at": _iso(datetime.now(timezone.utc)), "active": True,
        }
        await db.users.insert_one(user)
    else:
        await db.users.update_one({"email": email}, {"$set": {"picture": picture, "avatar_url": picture}})

    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    await db.user_sessions.insert_one({
        "session_token": session_token,
        "user_id": user["user_id"],
        "expires_at": _iso(expires_at),
        "created_at": _iso(datetime.now(timezone.utc)),
    })
    response.set_cookie(key="session_token", value=session_token,
                        httponly=True, secure=True, samesite="none",
                        max_age=7 * 24 * 3600, path="/")
    return {"user": _user_to_public(user)}


# ---------- Dashboard ----------
@api.get("/dashboard/stats")
async def dashboard_stats(request: Request):
    user = await get_current_user(request, db)
    uid = user["user_id"]
    today = date.today().isoformat()
    checked_today = await db.checkins.find_one({"user_id": uid, "date": today})
    prospects_count = await db.prospects.count_documents({"user_id": uid})
    prospects_won = await db.prospects.count_documents({"user_id": uid, "status": "won"})
    pending_followups = await db.followups.count_documents({"user_id": uid, "status": "pending"})
    total_attendance = await db.attendance.count_documents({"user_id": uid})
    active_challenges = await db.challenge_progress.count_documents({"user_id": uid, "completed": False})

    # weekly XP
    week_start = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
    weekly_xp_events = await db.xp_events.find({"user_id": uid, "created_at": {"$gte": week_start}}).to_list(1000)
    weekly_xp = sum(e.get("amount", 0) for e in weekly_xp_events)

    prog = xp_progress(user.get("xp", 0))
    return {
        "user": _user_to_public(user),
        **prog,
        "checked_in_today": bool(checked_today),
        "streak_current": user.get("streak_current", 0),
        "streak_longest": user.get("streak_longest", 0),
        "prospects_count": prospects_count,
        "prospects_won": prospects_won,
        "pending_followups": pending_followups,
        "total_attendance": total_attendance,
        "active_challenges": active_challenges,
        "weekly_xp": weekly_xp,
    }


@api.post("/checkins/daily")
async def daily_checkin(request: Request):
    user = await get_current_user(request, db)
    uid = user["user_id"]
    today = date.today()
    today_str = today.isoformat()
    exists = await db.checkins.find_one({"user_id": uid, "date": today_str})
    if exists:
        raise HTTPException(status_code=400, detail="Already checked in today")

    last = user.get("last_checkin_date")
    streak = user.get("streak_current", 0)
    longest = user.get("streak_longest", 0)
    if last:
        last_d = date.fromisoformat(last)
        delta_days = (today - last_d).days
        if delta_days == 1:
            streak += 1
        else:
            streak = 1
    else:
        streak = 1
    longest = max(longest, streak)

    await db.users.update_one({"user_id": uid}, {"$set": {
        "streak_current": streak, "streak_longest": longest, "last_checkin_date": today_str
    }})
    await db.checkins.insert_one({
        "checkin_id": str(uuid.uuid4()), "user_id": uid, "date": today_str,
        "xp_earned": XP_RULES["daily_checkin"],
        "created_at": _iso(datetime.now(timezone.utc)),
    })
    result = await award_xp(uid, XP_RULES["daily_checkin"], "daily_checkin")
    await _update_challenge_progress(uid, "checkins", 1)
    return {"streak_current": streak, "streak_longest": longest, **(result or {})}


# ---------- Prospects ----------
def _clean(doc):
    doc.pop("_id", None)
    return doc


@api.get("/prospects")
async def list_prospects(request: Request):
    user = await get_current_user(request, db)
    items = await db.prospects.find({"user_id": user["user_id"]}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return items


@api.post("/prospects")
async def add_prospect(payload: ProspectIn, request: Request):
    user = await get_current_user(request, db)
    doc = payload.model_dump()
    doc.update({
        "prospect_id": str(uuid.uuid4()),
        "user_id": user["user_id"],
        "created_at": _iso(datetime.now(timezone.utc)),
        "updated_at": _iso(datetime.now(timezone.utc)),
    })
    await db.prospects.insert_one(doc)
    xp = await award_xp(user["user_id"], XP_RULES["prospect_added"], "prospect_added")
    await _update_challenge_progress(user["user_id"], "prospects", 1)
    return {"prospect": _clean(doc), "xp": xp}


@api.patch("/prospects/{pid}")
async def update_prospect(pid: str, payload: ProspectUpdate, request: Request):
    user = await get_current_user(request, db)
    existing = await db.prospects.find_one({"prospect_id": pid, "user_id": user["user_id"]}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Prospect not found")
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    updates["updated_at"] = _iso(datetime.now(timezone.utc))
    await db.prospects.update_one({"prospect_id": pid}, {"$set": updates})
    xp_event = None
    if updates.get("status") == "won" and existing.get("status") != "won":
        xp_event = await award_xp(user["user_id"], XP_RULES["prospect_won"], "prospect_won")
    updated = await db.prospects.find_one({"prospect_id": pid}, {"_id": 0})
    return {"prospect": updated, "xp": xp_event}


@api.delete("/prospects/{pid}")
async def delete_prospect(pid: str, request: Request):
    user = await get_current_user(request, db)
    r = await db.prospects.delete_one({"prospect_id": pid, "user_id": user["user_id"]})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Prospect not found")
    return {"ok": True}


# ---------- Follow-ups ----------
@api.get("/followups")
async def list_followups(request: Request):
    user = await get_current_user(request, db)
    items = await db.followups.find({"user_id": user["user_id"]}, {"_id": 0}).sort("due_date", 1).to_list(500)
    return items


@api.post("/followups")
async def add_followup(payload: FollowUpIn, request: Request):
    user = await get_current_user(request, db)
    doc = payload.model_dump()
    doc.update({
        "followup_id": str(uuid.uuid4()),
        "user_id": user["user_id"],
        "status": "pending",
        "created_at": _iso(datetime.now(timezone.utc)),
    })
    await db.followups.insert_one(doc)
    return _clean(doc)


@api.patch("/followups/{fid}")
async def update_followup(fid: str, payload: FollowUpUpdate, request: Request):
    user = await get_current_user(request, db)
    existing = await db.followups.find_one({"followup_id": fid, "user_id": user["user_id"]}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Follow-up not found")
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    await db.followups.update_one({"followup_id": fid}, {"$set": updates})
    xp_event = None
    if updates.get("status") == "done" and existing.get("status") != "done":
        xp_event = await award_xp(user["user_id"], XP_RULES["followup_done"], "followup_done")
        await _update_challenge_progress(user["user_id"], "followups", 1)
    updated = await db.followups.find_one({"followup_id": fid}, {"_id": 0})
    return {"followup": updated, "xp": xp_event}


@api.delete("/followups/{fid}")
async def delete_followup(fid: str, request: Request):
    user = await get_current_user(request, db)
    r = await db.followups.delete_one({"followup_id": fid, "user_id": user["user_id"]})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}


# ---------- Attendance ----------
@api.get("/attendance")
async def list_attendance(request: Request):
    user = await get_current_user(request, db)
    items = await db.attendance.find({"user_id": user["user_id"]}, {"_id": 0}).sort("event_date", -1).to_list(500)
    return items


@api.post("/attendance")
async def add_attendance(payload: AttendanceIn, request: Request):
    user = await get_current_user(request, db)
    doc = payload.model_dump()
    doc.update({
        "attendance_id": str(uuid.uuid4()),
        "user_id": user["user_id"],
        "xp_earned": XP_RULES["attendance"],
        "created_at": _iso(datetime.now(timezone.utc)),
    })
    await db.attendance.insert_one(doc)
    xp = await award_xp(user["user_id"], XP_RULES["attendance"], "attendance")
    await _update_challenge_progress(user["user_id"], "attendance", 1)
    return {"attendance": _clean(doc), "xp": xp}


# ---------- Missions (GPS + Photo tracked field submissions) ----------
MAX_PHOTO_BYTES = 1_500_000  # ~1.5 MB base64


def _maps_url(lat: Optional[float], lng: Optional[float]) -> Optional[str]:
    if lat is None or lng is None:
        return None
    return f"https://www.google.com/maps?q={lat},{lng}"


@api.get("/missions")
async def list_missions(request: Request, limit: int = 100):
    user = await get_current_user(request, db)
    items = await db.missions.find(
        {"user_id": user["user_id"]},
        {"_id": 0},
    ).sort("created_at", -1).limit(limit).to_list(limit)
    return items


@api.post("/missions")
async def create_mission(payload: MissionIn, request: Request):
    user = await get_current_user(request, db)
    if payload.photo_data and len(payload.photo_data) > MAX_PHOTO_BYTES:
        raise HTTPException(status_code=413, detail="Photo too large (max ~1MB). Please retake.")
    doc = payload.model_dump()
    doc.update({
        "mission_id": str(uuid.uuid4()),
        "user_id": user["user_id"],
        "google_maps_url": _maps_url(payload.lat, payload.lng),
        "created_at": _iso(datetime.now(timezone.utc)),
        "updated_at": _iso(datetime.now(timezone.utc)),
    })
    await db.missions.insert_one(doc)
    xp_reason = "mission_converted" if payload.status == "converted" else "mission_logged"
    xp_amount = XP_RULES["mission_converted"] if payload.status == "converted" else XP_RULES["mission_logged"]
    xp = await award_xp(user["user_id"], xp_amount, xp_reason)
    await _update_challenge_progress(user["user_id"], "prospects", 1)
    return {"mission": _clean(doc), "xp": xp}


@api.patch("/missions/{mid}")
async def update_mission(mid: str, payload: MissionUpdate, request: Request):
    user = await get_current_user(request, db)
    existing = await db.missions.find_one({"mission_id": mid, "user_id": user["user_id"]}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Mission not found")
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    updates["updated_at"] = _iso(datetime.now(timezone.utc))
    await db.missions.update_one({"mission_id": mid}, {"$set": updates})
    xp_event = None
    if updates.get("status") == "converted" and existing.get("status") != "converted":
        xp_event = await award_xp(user["user_id"], XP_RULES["mission_converted"] - XP_RULES["mission_logged"], "mission_converted")
    updated = await db.missions.find_one({"mission_id": mid}, {"_id": 0})
    return {"mission": updated, "xp": xp_event}


@api.delete("/missions/{mid}")
async def delete_mission(mid: str, request: Request):
    user = await get_current_user(request, db)
    existing = await db.missions.find_one({"mission_id": mid, "user_id": user["user_id"]}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Mission not found")
    # Reverse the XP that was originally awarded for this mission.
    if existing.get("status") == "converted":
        reversal = -(XP_RULES["mission_converted"])
    else:
        reversal = -(XP_RULES["mission_logged"])
    await award_xp(user["user_id"], reversal, "mission_deleted")
    await db.missions.delete_one({"mission_id": mid})
    return {"ok": True, "xp_reversed": abs(reversal)}


# ---------- Admin Missions CRUD (Super Admin can manage any member's mission) ----------
@api.get("/admin/missions")
async def admin_list_all_missions(request: Request, limit: int = 500, status: Optional[str] = None):
    caller = await get_current_user(request, db)
    require_role(caller, ["super_admin"])
    q = {}
    if status:
        q["status"] = status
    items = await db.missions.find(q, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    uids = list({m["user_id"] for m in items})
    users = await db.users.find({"user_id": {"$in": uids}}, {"_id": 0, "user_id": 1, "name": 1, "team": 1, "avatar_url": 1, "picture": 1}).to_list(len(uids) or 1)
    umap = {u["user_id"]: u for u in users}
    for m in items:
        u = umap.get(m["user_id"], {})
        m["owner_name"] = u.get("name")
        m["owner_team"] = u.get("team")
        m["owner_avatar_url"] = u.get("avatar_url") or u.get("picture")
    return items


@api.patch("/admin/missions/{mid}")
async def admin_update_mission(mid: str, payload: MissionUpdate, request: Request):
    caller = await get_current_user(request, db)
    require_role(caller, ["super_admin"])
    existing = await db.missions.find_one({"mission_id": mid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Mission not found")
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    updates["updated_at"] = _iso(datetime.now(timezone.utc))
    updates["edited_by_admin"] = caller["user_id"]
    xp_delta = 0
    if updates.get("status") == "converted" and existing.get("status") != "converted":
        xp_delta = XP_RULES["mission_converted"] - XP_RULES["mission_logged"]
        await award_xp(existing["user_id"], xp_delta, "mission_converted_admin")
    elif updates.get("status") and updates["status"] != "converted" and existing.get("status") == "converted":
        # Downgrading a converted mission — reverse the conversion bonus
        xp_delta = -(XP_RULES["mission_converted"] - XP_RULES["mission_logged"])
        await award_xp(existing["user_id"], xp_delta, "mission_downgraded_admin")
    await db.missions.update_one({"mission_id": mid}, {"$set": updates})
    return {"ok": True, "xp_delta": xp_delta}


@api.delete("/admin/missions/{mid}")
async def admin_delete_mission(mid: str, request: Request):
    caller = await get_current_user(request, db)
    require_role(caller, ["super_admin"])
    existing = await db.missions.find_one({"mission_id": mid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Mission not found")
    # Reverse XP originally awarded
    if existing.get("status") == "converted":
        reversal = -(XP_RULES["mission_converted"])
    else:
        reversal = -(XP_RULES["mission_logged"])
    await award_xp(existing["user_id"], reversal, "mission_deleted_admin")
    await db.missions.delete_one({"mission_id": mid})
    return {"ok": True, "xp_reversed": abs(reversal), "owner": existing["user_id"]}


# ---------- Challenges ----------
@api.get("/challenges")
async def list_challenges(request: Request):
    user = await get_current_user(request, db)
    chs = await db.challenges.find({"end_date": {"$gte": date.today().isoformat()}}, {"_id": 0}).sort("start_date", 1).to_list(200)
    # Attach my progress
    progs = await db.challenge_progress.find({"user_id": user["user_id"]}, {"_id": 0}).to_list(500)
    prog_by = {p["challenge_id"]: p for p in progs}
    for c in chs:
        p = prog_by.get(c["challenge_id"])
        c["joined"] = bool(p)
        c["progress"] = p["progress"] if p else 0
        c["completed"] = p["completed"] if p else False
    return chs


@api.post("/challenges")
async def create_challenge(payload: ChallengeIn, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin", "team_leader"])
    doc = payload.model_dump()
    doc.update({
        "challenge_id": str(uuid.uuid4()),
        "created_by": user["user_id"],
        "created_at": _iso(datetime.now(timezone.utc)),
    })
    await db.challenges.insert_one(doc)
    return _clean(doc)


@api.post("/challenges/{cid}/join")
async def join_challenge(cid: str, request: Request):
    user = await get_current_user(request, db)
    ch = await db.challenges.find_one({"challenge_id": cid}, {"_id": 0})
    if not ch:
        raise HTTPException(status_code=404, detail="Challenge not found")
    existing = await db.challenge_progress.find_one({"challenge_id": cid, "user_id": user["user_id"]})
    if existing:
        return {"ok": True, "already_joined": True}
    await db.challenge_progress.insert_one({
        "progress_id": str(uuid.uuid4()),
        "challenge_id": cid,
        "user_id": user["user_id"],
        "progress": 0,
        "completed": False,
        "joined_at": _iso(datetime.now(timezone.utc)),
    })
    return {"ok": True}


async def _update_challenge_progress(user_id: str, goal_type: str, inc: int):
    today = date.today().isoformat()
    chs = await db.challenges.find({
        "goal_type": goal_type,
        "start_date": {"$lte": today},
        "end_date": {"$gte": today},
    }, {"_id": 0}).to_list(100)
    for ch in chs:
        prog = await db.challenge_progress.find_one({"challenge_id": ch["challenge_id"], "user_id": user_id})
        if not prog:
            continue
        if prog.get("completed"):
            continue
        new_progress = prog.get("progress", 0) + inc
        completed = new_progress >= ch["goal"]
        await db.challenge_progress.update_one(
            {"progress_id": prog["progress_id"]},
            {"$set": {"progress": new_progress, "completed": completed,
                      "completed_at": _iso(datetime.now(timezone.utc)) if completed else None}}
        )
        if completed:
            await award_xp(user_id, ch.get("xp_reward", 100), f"challenge:{ch['challenge_id']}")
            if ch.get("badge_reward"):
                await db.users.update_one({"user_id": user_id}, {"$addToSet": {"badges": ch["badge_reward"]}})


# ---------- Leaderboard ----------
@api.get("/leaderboard")
async def leaderboard(request: Request, scope: str = "all", limit: int = 50, team_id: Optional[str] = None):
    user = await get_current_user(request, db)
    # Team scoping: build member_id filter if requested
    member_ids = None
    if team_id:
        # Any authenticated user can request a team leaderboard (transparency)
        team = await _team_by_id(team_id)
        team_users = await db.users.find({"team_id": team["team_id"]}, {"_id": 0, "user_id": 1}).to_list(1000)
        member_ids = [t["user_id"] for t in team_users]
    if scope == "weekly":
        since = _iso(datetime.now(timezone.utc) - timedelta(days=7))
        match = {"created_at": {"$gte": since}}
        if member_ids is not None:
            match["user_id"] = {"$in": member_ids}
        pipeline = [
            {"$match": match},
            {"$group": {"_id": "$user_id", "xp": {"$sum": "$amount"}}},
            {"$sort": {"xp": -1}},
            {"$limit": limit},
        ]
        rows = await db.xp_events.aggregate(pipeline).to_list(limit)
        user_ids = [r["_id"] for r in rows]
        users = await db.users.find({"user_id": {"$in": user_ids}}, {"_id": 0, "password_hash": 0}).to_list(limit)
        u_by = {u["user_id"]: u for u in users}
        result = []
        for i, r in enumerate(rows):
            u = u_by.get(r["_id"])
            if not u:
                continue
            result.append({"rank": i + 1, "user_id": u["user_id"], "name": u["name"],
                           "avatar_url": u.get("avatar_url") or u.get("picture"),
                           "team": u.get("team"), "xp": r["xp"], "level": u.get("level", 1),
                           "streak_current": u.get("streak_current", 0)})
        return result
    elif scope == "monthly":
        since = _iso(datetime.now(timezone.utc) - timedelta(days=30))
        match = {"created_at": {"$gte": since}}
        if member_ids is not None:
            match["user_id"] = {"$in": member_ids}
        pipeline = [
            {"$match": match},
            {"$group": {"_id": "$user_id", "xp": {"$sum": "$amount"}}},
            {"$sort": {"xp": -1}},
            {"$limit": limit},
        ]
        rows = await db.xp_events.aggregate(pipeline).to_list(limit)
        user_ids = [r["_id"] for r in rows]
        users = await db.users.find({"user_id": {"$in": user_ids}}, {"_id": 0, "password_hash": 0}).to_list(limit)
        u_by = {u["user_id"]: u for u in users}
        result = []
        for i, r in enumerate(rows):
            u = u_by.get(r["_id"])
            if not u:
                continue
            result.append({"rank": i + 1, "user_id": u["user_id"], "name": u["name"],
                           "avatar_url": u.get("avatar_url") or u.get("picture"),
                           "team": u.get("team"), "xp": r["xp"], "level": u.get("level", 1),
                           "streak_current": u.get("streak_current", 0)})
        return result
    else:
        query = {}
        if member_ids is not None:
            query["user_id"] = {"$in": member_ids}
        users = await db.users.find(query, {"_id": 0, "password_hash": 0}).sort("xp", -1).limit(limit).to_list(limit)
        return [{"rank": i + 1, "user_id": u["user_id"], "name": u["name"],
                 "avatar_url": u.get("avatar_url") or u.get("picture"),
                 "team": u.get("team"), "xp": u.get("xp", 0), "level": u.get("level", 1),
                 "streak_current": u.get("streak_current", 0)}
                for i, u in enumerate(users)]


@api.get("/leaderboard/teams")
async def team_leaderboard(request: Request):
    await get_current_user(request, db)
    pipeline = [
        {"$group": {"_id": "$team", "xp": {"$sum": "$xp"}, "members": {"$sum": 1}}},
        {"$sort": {"xp": -1}},
    ]
    rows = await db.users.aggregate(pipeline).to_list(50)
    return [{"rank": i + 1, "team": r["_id"] or "Unassigned", "xp": r["xp"], "members": r["members"]}
            for i, r in enumerate(rows)]


# ---------- Badges ----------
@api.get("/badges")
async def list_badges(request: Request):
    user = await get_current_user(request, db)
    owned = set(user.get("badges", []))
    return [{**b, "unlocked": b["key"] in owned} for b in BADGE_CATALOG]


# ---------- Teams ----------
async def _team_by_id(team_id: str):
    t = await db.teams.find_one({"team_id": team_id}, {"_id": 0})
    if not t:
        raise HTTPException(status_code=404, detail="Team not found")
    return t


async def _my_team_or_403(user: dict) -> dict:
    """Return the team the given team_leader leads, or 403."""
    require_role(user, ["team_leader", "super_admin"])
    if user["role"] == "super_admin":
        raise HTTPException(status_code=400, detail="Super admin has no single team")
    team = await db.teams.find_one({"leader_id": user["user_id"]}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="You do not lead a team yet")
    return team


@api.get("/teams")
async def list_teams(request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    teams = await db.teams.find({}, {"_id": 0}).sort("name", 1).to_list(200)
    # Attach member counts + leader info
    for t in teams:
        t["member_count"] = await db.users.count_documents({"team_id": t["team_id"]})
        if t.get("leader_id"):
            leader = await db.users.find_one({"user_id": t["leader_id"]}, {"_id": 0, "password_hash": 0})
            t["leader"] = {"user_id": leader["user_id"], "name": leader["name"], "email": leader["email"]} if leader else None
        else:
            t["leader"] = None
    return teams


@api.post("/teams")
async def create_team(payload: TeamIn, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    exists = await db.teams.find_one({"name": payload.name})
    if exists:
        raise HTTPException(status_code=400, detail="Team name already exists")
    doc = {
        "team_id": str(uuid.uuid4()),
        "name": payload.name,
        "leader_id": payload.leader_id,
        "created_at": _iso(datetime.now(timezone.utc)),
    }
    await db.teams.insert_one(doc)
    if payload.leader_id:
        await db.users.update_one(
            {"user_id": payload.leader_id},
            {"$set": {"role": "team_leader", "team_id": doc["team_id"], "team": payload.name}},
        )
    return _clean(doc)


@api.patch("/teams/{team_id}")
async def update_team(team_id: str, payload: TeamUpdate, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    team = await _team_by_id(team_id)
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not updates:
        return team
    if "name" in updates and updates["name"] != team["name"]:
        clash = await db.teams.find_one({"name": updates["name"]})
        if clash:
            raise HTTPException(status_code=400, detail="Team name already exists")
        # Propagate name change to users
        await db.users.update_many({"team_id": team_id}, {"$set": {"team": updates["name"]}})
    if "leader_id" in updates:
        new_leader = updates["leader_id"]
        # Demote previous leader if any
        if team.get("leader_id") and team["leader_id"] != new_leader:
            prev = await db.users.find_one({"user_id": team["leader_id"]}, {"_id": 0})
            if prev and prev.get("role") == "team_leader":
                await db.users.update_one({"user_id": team["leader_id"]}, {"$set": {"role": "member"}})
        if new_leader:
            await db.users.update_one(
                {"user_id": new_leader},
                {"$set": {"role": "team_leader", "team_id": team_id, "team": updates.get("name", team["name"])}},
            )
    await db.teams.update_one({"team_id": team_id}, {"$set": updates})
    return await _team_by_id(team_id)


@api.delete("/teams/{team_id}")
async def delete_team(team_id: str, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    team = await _team_by_id(team_id)
    # Un-assign all members
    await db.users.update_many({"team_id": team_id}, {"$set": {"team_id": None, "team": None}})
    # Demote leader
    if team.get("leader_id"):
        await db.users.update_one(
            {"user_id": team["leader_id"], "role": "team_leader"},
            {"$set": {"role": "member"}},
        )
    await db.teams.delete_one({"team_id": team_id})
    return {"ok": True}


@api.post("/teams/{team_id}/assign")
async def assign_to_team(team_id: str, payload: TeamAssign, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    team = await _team_by_id(team_id)
    target = await db.users.find_one({"user_id": payload.user_id}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    updates = {"team_id": team_id, "team": team["name"]}
    if payload.is_leader:
        updates["role"] = "team_leader"
        # Demote previous leader if different
        if team.get("leader_id") and team["leader_id"] != payload.user_id:
            await db.users.update_one(
                {"user_id": team["leader_id"], "role": "team_leader"},
                {"$set": {"role": "member"}},
            )
        await db.teams.update_one({"team_id": team_id}, {"$set": {"leader_id": payload.user_id}})
    await db.users.update_one({"user_id": payload.user_id}, {"$set": updates})
    return {"ok": True}


@api.delete("/teams/{team_id}/members/{user_id}")
async def remove_from_team(team_id: str, user_id: str, request: Request):
    admin = await get_current_user(request, db)
    require_role(admin, ["super_admin"])
    team = await _team_by_id(team_id)
    target = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target or target.get("team_id") != team_id:
        raise HTTPException(status_code=404, detail="User is not on this team")
    updates = {"team_id": None, "team": None}
    if team.get("leader_id") == user_id:
        updates["role"] = "member"
        await db.teams.update_one({"team_id": team_id}, {"$set": {"leader_id": None}})
    await db.users.update_one({"user_id": user_id}, {"$set": updates})
    return {"ok": True}


@api.get("/my-team")
async def my_team(request: Request):
    user = await get_current_user(request, db)
    team = await _my_team_or_403(user)
    members = await db.users.find(
        {"team_id": team["team_id"]}, {"_id": 0, "password_hash": 0}
    ).sort("xp", -1).to_list(500)
    return {"team": team, "members": members}


# ---------- Reports ----------
async def _user_report(uid: str):
    checkins = await db.checkins.count_documents({"user_id": uid})
    prospects = await db.prospects.count_documents({"user_id": uid})
    won = await db.prospects.count_documents({"user_id": uid, "status": "won"})
    lost = await db.prospects.count_documents({"user_id": uid, "status": "lost"})
    followups_done = await db.followups.count_documents({"user_id": uid, "status": "done"})
    followups_pending = await db.followups.count_documents({"user_id": uid, "status": "pending"})
    attendance = await db.attendance.count_documents({"user_id": uid})
    since = _iso(datetime.now(timezone.utc) - timedelta(days=30))
    xp_30d = await db.xp_events.aggregate([
        {"$match": {"user_id": uid, "created_at": {"$gte": since}}},
        {"$group": {"_id": None, "xp": {"$sum": "$amount"}}},
    ]).to_list(1)
    # Timeline: last 14 days daily XP
    days = 14
    day_start = date.today() - timedelta(days=days - 1)
    timeline = []
    for i in range(days):
        d = day_start + timedelta(days=i)
        d_start = _iso(datetime.combine(d, datetime.min.time()).replace(tzinfo=timezone.utc))
        d_end = _iso(datetime.combine(d + timedelta(days=1), datetime.min.time()).replace(tzinfo=timezone.utc))
        agg = await db.xp_events.aggregate([
            {"$match": {"user_id": uid, "created_at": {"$gte": d_start, "$lt": d_end}}},
            {"$group": {"_id": None, "xp": {"$sum": "$amount"}}},
        ]).to_list(1)
        timeline.append({"date": d.isoformat(), "xp": agg[0]["xp"] if agg else 0})
    return {
        "checkins": checkins, "prospects": prospects, "won": won, "lost": lost,
        "conversion_rate": round((won / prospects) * 100, 1) if prospects else 0.0,
        "followups_done": followups_done, "followups_pending": followups_pending,
        "attendance": attendance,
        "xp_30d": xp_30d[0]["xp"] if xp_30d else 0,
        "timeline": timeline,
    }


@api.get("/reports/me")
async def report_me(request: Request):
    user = await get_current_user(request, db)
    stats = await _user_report(user["user_id"])
    return {"user": user, **stats}


@api.get("/reports/team")
async def report_team(request: Request, team_id: Optional[str] = None):
    """Team-scoped report. team_leader: their team only. super_admin: any team via team_id."""
    user = await get_current_user(request, db)
    require_role(user, ["team_leader", "super_admin"])
    if user["role"] == "team_leader":
        team = await _my_team_or_403(user)
    else:
        if not team_id:
            raise HTTPException(status_code=400, detail="team_id required for super_admin")
        team = await _team_by_id(team_id)

    members = await db.users.find(
        {"team_id": team["team_id"]}, {"_id": 0, "password_hash": 0}
    ).sort("xp", -1).to_list(500)
    member_ids = [m["user_id"] for m in members]

    total_prospects = await db.prospects.count_documents({"user_id": {"$in": member_ids}})
    total_won = await db.prospects.count_documents({"user_id": {"$in": member_ids}, "status": "won"})
    total_checkins = await db.checkins.count_documents({"user_id": {"$in": member_ids}})
    total_attendance = await db.attendance.count_documents({"user_id": {"$in": member_ids}})
    total_followups_done = await db.followups.count_documents({"user_id": {"$in": member_ids}, "status": "done"})
    total_xp = sum(m.get("xp", 0) for m in members)
    active_today = await db.checkins.count_documents({
        "user_id": {"$in": member_ids}, "date": date.today().isoformat()
    })

    # Per-member breakdown
    member_stats = []
    for m in members:
        m_prospects = await db.prospects.count_documents({"user_id": m["user_id"]})
        m_won = await db.prospects.count_documents({"user_id": m["user_id"], "status": "won"})
        m_followups = await db.followups.count_documents({"user_id": m["user_id"], "status": "done"})
        m_attendance = await db.attendance.count_documents({"user_id": m["user_id"]})
        member_stats.append({
            "user_id": m["user_id"], "name": m["name"], "email": m["email"],
            "role": m["role"], "xp": m.get("xp", 0), "level": m.get("level", 1),
            "streak_current": m.get("streak_current", 0),
            "avatar_url": m.get("avatar_url") or m.get("picture"),
            "position_badges": m.get("position_badges", []),
            "prospects": m_prospects, "won": m_won,
            "followups_done": m_followups, "attendance": m_attendance,
        })

    return {
        "team": team,
        "totals": {
            "members": len(members), "prospects": total_prospects, "won": total_won,
            "checkins": total_checkins, "attendance": total_attendance,
            "followups_done": total_followups_done, "xp": total_xp,
            "active_today": active_today,
            "conversion_rate": round((total_won / total_prospects) * 100, 1) if total_prospects else 0.0,
        },
        "members": member_stats,
    }


@api.get("/reports/global")
async def report_global(request: Request):
    """Super admin only — full org-wide report grouped by team."""
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])

    teams = await db.teams.find({}, {"_id": 0}).to_list(200)
    per_team = []
    for t in teams:
        members = await db.users.find({"team_id": t["team_id"]}, {"_id": 0, "password_hash": 0}).to_list(500)
        mids = [m["user_id"] for m in members]
        prospects = await db.prospects.count_documents({"user_id": {"$in": mids}}) if mids else 0
        won = await db.prospects.count_documents({"user_id": {"$in": mids}, "status": "won"}) if mids else 0
        xp = sum(m.get("xp", 0) for m in members)
        per_team.append({
            "team_id": t["team_id"], "name": t["name"], "members": len(members),
            "prospects": prospects, "won": won, "xp": xp,
            "conversion_rate": round((won / prospects) * 100, 1) if prospects else 0.0,
        })

    # Org totals
    total_users = await db.users.count_documents({})
    total_prospects = await db.prospects.count_documents({})
    total_won = await db.prospects.count_documents({"status": "won"})
    total_teams = await db.teams.count_documents({})

    return {
        "totals": {
            "users": total_users, "teams": total_teams,
            "prospects": total_prospects, "won": total_won,
            "conversion_rate": round((total_won / total_prospects) * 100, 1) if total_prospects else 0.0,
        },
        "teams": sorted(per_team, key=lambda t: t["xp"], reverse=True),
    }


# ---------- Admin ----------
@api.get("/admin/users")
async def admin_users(request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin", "team_leader"])
    if user["role"] == "team_leader":
        team = await _my_team_or_403(user)
        users = await db.users.find(
            {"team_id": team["team_id"]}, {"_id": 0, "password_hash": 0}
        ).sort("xp", -1).to_list(1000)
    else:
        users = await db.users.find({}, {"_id": 0, "password_hash": 0}).sort("xp", -1).to_list(1000)
    return users


@api.patch("/admin/users/{uid}/role")
async def admin_update_role(uid: str, payload: RoleUpdate, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    r = await db.users.update_one({"user_id": uid}, {"$set": {"role": payload.role}})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"ok": True}


@api.get("/admin/analytics")
async def admin_analytics(request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin", "team_leader"])

    # Team-leader scoped
    if user["role"] == "team_leader":
        team = await _my_team_or_403(user)
        member_ids = [
            m["user_id"] async for m in db.users.find({"team_id": team["team_id"]}, {"_id": 0, "user_id": 1})
        ]
        match = {"user_id": {"$in": member_ids}} if member_ids else {"user_id": {"$in": []}}
        total_users = len(member_ids)
        total_prospects = await db.prospects.count_documents(match)
        total_won = await db.prospects.count_documents({**match, "status": "won"})
        total_attendance = await db.attendance.count_documents(match)
        total_checkins = await db.checkins.count_documents(match)
        active_today = await db.checkins.count_documents({**match, "date": date.today().isoformat()})
        since = _iso(datetime.now(timezone.utc) - timedelta(days=7))
        weekly = await db.xp_events.aggregate([
            {"$match": {**match, "created_at": {"$gte": since}}},
            {"$group": {"_id": None, "xp": {"$sum": "$amount"}}},
        ]).to_list(1)
    else:
        total_users = await db.users.count_documents({})
        total_prospects = await db.prospects.count_documents({})
        total_won = await db.prospects.count_documents({"status": "won"})
        total_attendance = await db.attendance.count_documents({})
        total_checkins = await db.checkins.count_documents({})
        active_today = await db.checkins.count_documents({"date": date.today().isoformat()})
        since = _iso(datetime.now(timezone.utc) - timedelta(days=7))
        weekly = await db.xp_events.aggregate([
            {"$match": {"created_at": {"$gte": since}}},
            {"$group": {"_id": None, "xp": {"$sum": "$amount"}}},
        ]).to_list(1)

    weekly_xp = weekly[0]["xp"] if weekly else 0
    return {
        "scope": "team" if user["role"] == "team_leader" else "global",
        "total_users": total_users, "total_prospects": total_prospects, "total_won": total_won,
        "total_attendance": total_attendance, "total_checkins": total_checkins,
        "active_today": active_today, "weekly_xp": weekly_xp,
        "conversion_rate": round((total_won / total_prospects) * 100, 1) if total_prospects else 0.0,
    }


@api.get("/admin/dashboard-widgets")
async def admin_dashboard_widgets(request: Request):
    """Extended admin dashboard: missions today, pending tasks, top individual, top team,
    season champion, upcoming birthdays / anniversaries."""
    user = await get_current_user(request, db)
    require_role(user, ["super_admin", "team_leader"])

    today = date.today()
    today_iso = today.isoformat()
    today_start_iso = today_iso + "T00:00:00+00:00"

    # Scope
    if user["role"] == "team_leader":
        team = await _my_team_or_403(user)
        member_ids = [m["user_id"] async for m in db.users.find({"team_id": team["team_id"]}, {"_id": 0, "user_id": 1})]
        user_filter = {"user_id": {"$in": member_ids}} if member_ids else {"user_id": {"$in": []}}
        assignee_filter = {"assigned_to": {"$in": member_ids}} if member_ids else {"assigned_to": {"$in": []}}
        users_all = await db.users.find({"team_id": team["team_id"]}, {"_id": 0, "password_hash": 0}).to_list(1000)
    else:
        user_filter = {}
        assignee_filter = {}
        users_all = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(5000)

    # Missions today (created today)
    missions_today = await db.missions.count_documents({
        **user_filter, "created_at": {"$gte": today_start_iso},
    })
    missions_converted_today = await db.missions.count_documents({
        **user_filter, "status": "converted", "updated_at": {"$gte": today_start_iso},
    })

    # Pending tasks
    pending_tasks = await db.tasks.count_documents({**assignee_filter, "status": "pending"})
    overdue_tasks = await db.tasks.count_documents({
        **assignee_filter, "status": "pending", "due_date": {"$lt": today_iso},
    })

    # Top individual (by all-time XP)
    top_individual = None
    if users_all:
        top_u = max(users_all, key=lambda u: u.get("xp", 0))
        if top_u.get("xp", 0) > 0:
            top_individual = {
                "user_id": top_u["user_id"], "name": top_u["name"],
                "team": top_u.get("team"), "xp": top_u.get("xp", 0),
                "level": top_u.get("level", 1),
                "position_badges": top_u.get("position_badges", []),
            }

    # Top team (by XP sum)
    top_team = None
    teams = await db.teams.find({}, {"_id": 0}).to_list(200)
    team_scores = []
    for t in teams:
        members = [u for u in users_all if u.get("team_id") == t["team_id"]] if user["role"] == "team_leader" \
                  else await db.users.find({"team_id": t["team_id"]}, {"_id": 0, "xp": 1}).to_list(1000)
        xp = sum(m.get("xp", 0) for m in members)
        if xp > 0:
            team_scores.append({"team_id": t["team_id"], "name": t["name"], "xp": xp, "members": len(members)})
    if team_scores:
        top_team = max(team_scores, key=lambda t: t["xp"])

    # Season champion — top XP earner in currently active season
    active_season = await _current_active_season()
    season_champion = None
    if active_season:
        start_iso, end_iso = _season_iso_range(active_season)
        pipeline = [
            {"$match": {"created_at": {"$gte": start_iso, "$lte": end_iso}}},
        ]
        if user["role"] == "team_leader" and member_ids:
            pipeline[0]["$match"]["user_id"] = {"$in": member_ids}
        pipeline += [
            {"$group": {"_id": "$user_id", "xp": {"$sum": "$amount"}}},
            {"$sort": {"xp": -1}},
            {"$limit": 1},
        ]
        rows = await db.xp_events.aggregate(pipeline).to_list(1)
        if rows:
            u = await db.users.find_one({"user_id": rows[0]["_id"]}, {"_id": 0, "password_hash": 0})
            if u:
                season_champion = {
                    "user_id": u["user_id"], "name": u["name"], "team": u.get("team"),
                    "xp_in_season": rows[0]["xp"], "level": u.get("level", 1),
                    "position_badges": u.get("position_badges", []),
                    "season_name": active_season["name"],
                }

    # Upcoming celebrations (next 14 days including today)
    upcoming_birthdays = []
    upcoming_anniversaries = []
    for u in users_all:
        for field, out in [("dob", upcoming_birthdays), ("anniversary_date", upcoming_anniversaries)]:
            v = u.get(field)
            if not v:
                continue
            try:
                d = date.fromisoformat(v)
            except Exception:
                continue
            if field == "anniversary_date" and u.get("marital_status") != "married":
                continue
            # Find next occurrence
            try:
                this_year = d.replace(year=today.year)
            except ValueError:
                # Handle Feb 29 → Feb 28
                this_year = d.replace(year=today.year, day=28)
            if this_year < today:
                try:
                    this_year = this_year.replace(year=today.year + 1)
                except ValueError:
                    this_year = this_year.replace(year=today.year + 1, day=28)
            days_until = (this_year - today).days
            if days_until <= 14:
                out.append({
                    "user_id": u["user_id"], "name": u["name"],
                    "team": u.get("team"), "date": this_year.isoformat(),
                    "days_until": days_until,
                    "avatar_url": u.get("avatar_url") or u.get("avatar_photo"),
                })
    upcoming_birthdays.sort(key=lambda x: x["days_until"])
    upcoming_anniversaries.sort(key=lambda x: x["days_until"])

    return {
        "missions_today": missions_today,
        "missions_converted_today": missions_converted_today,
        "pending_tasks": pending_tasks,
        "overdue_tasks": overdue_tasks,
        "top_individual": top_individual,
        "top_team": top_team,
        "season_champion": season_champion,
        "upcoming_birthdays": upcoming_birthdays[:5],
        "upcoming_anniversaries": upcoming_anniversaries[:5],
    }


# ---------- Weekly Attendance / Seasons / Tasks ----------
try:
    from zoneinfo import ZoneInfo
    LOCAL_TZ = ZoneInfo("Asia/Kolkata")
except Exception:
    LOCAL_TZ = timezone.utc

LOCK_HOUR = 8  # 8 AM IST default cutoff
SATURDAY_LOCK_HOUR = 22  # 10 PM IST — Spartans Team Meeting stays open all day


def _lock_hour_for_weekday(wd: int) -> int:
    return SATURDAY_LOCK_HOUR if wd == 5 else LOCK_HOUR  # Mon=0 ... Sat=5


def _is_locked(event_date_str: str) -> bool:
    """True if attendance for event_date is closed. Saturday (Spartans Team Meeting) locks at 22:00; others at 08:00."""
    try:
        d = date.fromisoformat(event_date_str)
    except Exception:
        return True
    hour = _lock_hour_for_weekday(d.weekday())
    cutoff = datetime.combine(d, datetime.min.time().replace(hour=hour), tzinfo=LOCAL_TZ)
    return datetime.now(LOCAL_TZ) >= cutoff


def _weekday_name(w: int) -> str:
    return ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"][w]


def _dates_for_week(anchor: date):
    """Return Mon..Sun dates for the week containing anchor (Mon=0)."""
    monday = anchor - timedelta(days=anchor.weekday())
    return [monday + timedelta(days=i) for i in range(7)]


# --- Weekly Events (admin managed) ---
@api.get("/weekly-events")
async def list_weekly_events(request: Request):
    await get_current_user(request, db)
    events = await db.weekly_events.find({"active": True}, {"_id": 0}).sort("weekday", 1).to_list(100)
    for e in events:
        e["weekday_name"] = _weekday_name(e["weekday"])
    return events


@api.post("/weekly-events")
async def create_weekly_event(payload: WeeklyEventIn, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])

    if not payload.clubs:
        raise HTTPException(status_code=400, detail="Select at least one club")

    doc = payload.model_dump()
    doc.update({
        "event_id": str(uuid.uuid4()),
        "is_believer": "believer" in payload.clubs,
        "created_at": _iso(datetime.now(timezone.utc)),
    })

    await db.weekly_events.insert_one(doc)
    return _clean(doc)
    
@api.delete("/weekly-events/{event_id}")
async def delete_weekly_event(event_id: str, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])

    r = await db.weekly_events.update_one(
        {"event_id": event_id},
        {"$set": {"active": False, "deleted_at": _iso(datetime.now(timezone.utc))}}
    )

    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="Event not found")

    return {"ok": True}

@api.patch("/weekly-events/{event_id}/complete")
async def complete_weekly_event(event_id: str, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])

    r = await db.weekly_events.update_one(
        {"event_id": event_id},
        {"$set": {"completed": True, "completed_at": _iso(datetime.now(timezone.utc))}}
    )

    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="Event not found")

    return {"ok": True}


@api.get("/event-attendance/week")
async def week_attendance(request: Request, week_of: Optional[str] = None):
    """Returns attendance events.

    - Default: shows upcoming 30 days.
    - Future meetings are visible but locked.
    - Attendance can be marked only on meeting day.
    - Decider users will not see believer/season attendance.
    """
    user = await get_current_user(request, db)

    today = date.today()
    is_admin = user["role"] == "super_admin"

    if week_of:
        anchor = date.fromisoformat(week_of)
        week_dates = _dates_for_week(anchor)
        period_start = week_dates[0]
        period_end = week_dates[6]
    else:
        period_start = today
        period_end = today + timedelta(days=30)
event_filter = {"active": True, "completed": {"$ne": True}}

if not is_admin:
    user_club = str(user.get("club_type", "")).lower()
    event_filter["$or"] = [
        {"clubs": {"$in": [user_club]}},
        {"clubs": {"$exists": False}},
    ]
    
    if user_club == "decider":
        event_filter["is_believer"] = False

    events = await db.weekly_events.find(
        event_filter,
        {"_id": 0}
    ).sort("weekday", 1).to_list(100)

    occurrences = []
    period_dates = [
        period_start + timedelta(days=i)
        for i in range((period_end - period_start).days + 1)
    ]

    for e in events:
        event_weekday = int(e["weekday"])

        for occ_date in period_dates:
            if occ_date.weekday() != event_weekday:
                continue

            occ_date_str = occ_date.isoformat()

            mark = await db.event_attendance.find_one(
                {
                    "user_id": user["user_id"],
                    "event_id": e["event_id"],
                    "event_date": occ_date_str,
                },
                {"_id": 0},
            )

            occurrences.append({
                "event_id": e["event_id"],
                "name": e["name"],
                "weekday": e["weekday"],
                "weekday_name": _weekday_name(e["weekday"]),
                "is_believer": e.get("is_believer", False),
                "club_type": e.get("club_type", "all"),
"repeat_type": e.get("repeat_type", "weekly"),
"open_time": e.get("open_time", "06:00"),
"lock_time": e.get("lock_time", "22:00"),
"completed": e.get("completed", False),
                "event_date": occ_date_str,
                "status": mark["status"] if mark else None,
                "locked": True if occ_date > today else _is_locked(occ_date_str),
                "is_future": occ_date > today,
                "is_today": occ_date == today,
                "marked_at": mark.get("updated_at") if mark else None,
            })

    occurrences.sort(key=lambda x: x["event_date"])

    return {
        "week_start": period_start.isoformat(),
        "week_end": period_end.isoformat(),
        "occurrences": occurrences,
    }

@api.post("/event-attendance/mark")
async def mark_attendance(payload: EventAttendanceMark, request: Request):
    user = await get_current_user(request, db)
    # Only allow marking on the exact meeting day. Future dates are blocked; past dates are locked.
    try:
        ev_date = date.fromisoformat(payload.event_date)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid event_date")
    today = date.today()
    if ev_date > today:
        raise HTTPException(status_code=403, detail="Cannot mark attendance for a future meeting")
    if ev_date < today:
        raise HTTPException(status_code=403, detail="Past meeting attendance is locked (read-only history)")
    if _is_locked(payload.event_date):
        ev_wd = date.fromisoformat(payload.event_date).weekday()
        hh = _lock_hour_for_weekday(ev_wd)
        raise HTTPException(status_code=403, detail=f"Attendance locked. Cutoff is {hh:02d}:00 IST on {payload.event_date}.")
    # Validate event exists
    ev = await db.weekly_events.find_one({"event_id": payload.event_id, "active": True}, {"_id": 0})
    if not ev:
        raise HTTPException(status_code=404, detail="Event not found")
    # Auto-detect season if not passed
    season_id = payload.season_id
    if not season_id:
        s = await db.seasons.find_one({
            "start_date": {"$lte": payload.event_date},
            "end_date": {"$gte": payload.event_date},
            "$or": [{"is_believer": ev.get("is_believer", False)}, {"is_believer": False}],
        }, {"_id": 0})
        if s:
            season_id = s["season_id"]
    now = _iso(datetime.now(timezone.utc))
    key = {"user_id": user["user_id"], "event_id": payload.event_id, "event_date": payload.event_date}
    await db.event_attendance.update_one(
        key,
        {"$set": {**key, "status": payload.status, "season_id": season_id, "updated_at": now},
         "$setOnInsert": {"attendance_id": str(uuid.uuid4()), "created_at": now}},
        upsert=True,
    )
    ev_hour = _lock_hour_for_weekday(ev_date.weekday())
    return {"ok": True, "locks_at": f"{payload.event_date} {ev_hour:02d}:00 IST"}


# ---------- Admin / Team Leader mark-for-member attendance ----------
class MarkForMemberIn(BaseModel):
    user_id: str
    event_id: str
    event_date: str
    status: Literal["present", "absent", "excused"]
    season_id: Optional[str] = None


@api.post("/event-attendance/mark-for-member")
async def mark_attendance_for_member(payload: MarkForMemberIn, request: Request):
    """Super Admin can mark attendance for any member.
    Team Leaders can mark attendance for members of their own team."""
    caller = await get_current_user(request, db)
    require_role(caller, ["super_admin", "team_leader"])
    target = await db.users.find_one({"user_id": payload.user_id}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="Target user not found")
    if caller["role"] == "team_leader":
        team = await _my_team_or_403(caller)
        if target.get("team_id") != team["team_id"]:
            raise HTTPException(status_code=403, detail="You can only mark attendance for your own team")
    try:
        ev_date = date.fromisoformat(payload.event_date)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid event_date")
    today = date.today()
    if ev_date > today:
        raise HTTPException(status_code=403, detail="Cannot mark attendance for a future meeting")
    # Super admin bypass for retroactive past-date marks; team_leader still respects lock.
    if caller["role"] != "super_admin":
        if ev_date < today:
            raise HTTPException(status_code=403, detail="Past meeting attendance is locked")
        if _is_locked(payload.event_date):
            hh = _lock_hour_for_weekday(ev_date.weekday())
            raise HTTPException(status_code=403, detail=f"Attendance locked. Cutoff is {hh:02d}:00 IST on {payload.event_date}.")
    ev = await db.weekly_events.find_one({"event_id": payload.event_id, "active": True}, {"_id": 0})
    if not ev:
        raise HTTPException(status_code=404, detail="Event not found")
    season_id = payload.season_id
    if not season_id:
        s = await db.seasons.find_one({
            "start_date": {"$lte": payload.event_date},
            "end_date": {"$gte": payload.event_date},
        }, {"_id": 0})
        if s:
            season_id = s["season_id"]
    now = _iso(datetime.now(timezone.utc))
    key = {"user_id": payload.user_id, "event_id": payload.event_id, "event_date": payload.event_date}
    await db.event_attendance.update_one(
        key,
        {"$set": {**key, "status": payload.status, "season_id": season_id, "updated_at": now,
                  "marked_by": caller["user_id"], "marked_by_role": caller["role"]},
         "$setOnInsert": {"attendance_id": str(uuid.uuid4()), "created_at": now}},
        upsert=True,
    )
    return {"ok": True, "target": payload.user_id, "status": payload.status}


@api.get("/event-attendance/team-week")
async def team_week_attendance(request: Request, week_of: Optional[str] = None):
    """Return team-scoped attendance grid for admin & team-leader:
    columns = occurrences this week, rows = members, cells = mark status."""
    caller = await get_current_user(request, db)
    require_role(caller, ["super_admin", "team_leader"])
    anchor = date.fromisoformat(week_of) if week_of else date.today()
    week_dates = _dates_for_week(anchor)
    events = await db.weekly_events.find({"active": True}, {"_id": 0}).sort("weekday", 1).to_list(100)
    if caller["role"] == "team_leader":
        team = await _my_team_or_403(caller)
        members = await db.users.find({"team_id": team["team_id"]}, {"_id": 0, "password_hash": 0}).to_list(500)
    else:
        members = await db.users.find({}, {"_id": 0, "password_hash": 0}).sort("name", 1).to_list(5000)
    # Build occurrences (all — admin/leader can view full week grid)
    occurrences = [{
        "event_id": e["event_id"], "name": e["name"], "weekday": e["weekday"],
        "weekday_name": _weekday_name(e["weekday"]),
        "event_date": week_dates[e["weekday"]].isoformat(),
        "is_believer": e.get("is_believer", False),
        "is_locked": _is_locked(week_dates[e["weekday"]].isoformat()),
        "lock_hour": _lock_hour_for_weekday(e["weekday"]),
    } for e in events]
    # Fetch existing marks for these members + dates
    dates_iso = [o["event_date"] for o in occurrences]
    marks = await db.event_attendance.find({
        "user_id": {"$in": [m["user_id"] for m in members]},
        "event_date": {"$in": dates_iso},
    }, {"_id": 0}).to_list(20000)
    marks_map = {(m["user_id"], m["event_id"], m["event_date"]): m for m in marks}
    grid = []
    for m in members:
        row = {
            "user_id": m["user_id"], "name": m["name"], "team": m.get("team"),
            "avatar_url": m.get("avatar_url") or m.get("picture"),
            "position_badges": m.get("position_badges", []),
            "role": m.get("role"),
            "marks": {},
        }
        for o in occurrences:
            mk = marks_map.get((m["user_id"], o["event_id"], o["event_date"]))
            row["marks"][o["event_id"]] = {
                "status": mk["status"] if mk else None,
                "marked_by_role": mk.get("marked_by_role") if mk else None,
            }
        grid.append(row)
    return {"occurrences": occurrences, "grid": grid}


# --- Seasons ---
def _season_range_dates(s: dict):
    start = date.fromisoformat(s["start_date"])
    end = date.fromisoformat(s["end_date"])
    return start, end


async def _compute_report(user_id: str, season: dict, believer_only: bool = False):
    start, end = _season_range_dates(season)
    days = (end - start).days + 1
    # Get relevant weekly events
    ev_query = {"active": True}
    if believer_only:
        ev_query["is_believer"] = True
    events = await db.weekly_events.find(ev_query, {"_id": 0}).to_list(50)
    ev_by_wd = {}
    for e in events:
        ev_by_wd.setdefault(e["weekday"], []).append(e)
    total = 0
    per_event = {e["event_id"]: {"event_id": e["event_id"], "name": e["name"], "weekday": e["weekday"],
                                  "present": 0, "absent": 0, "na": 0, "unmarked": 0, "total": 0}
                 for e in events}
    for i in range(days):
        d = start + timedelta(days=i)
        for e in ev_by_wd.get(d.weekday(), []):
            total += 1
            per_event[e["event_id"]]["total"] += 1
            occ_date = d.isoformat()
            mark = await db.event_attendance.find_one(
                {"user_id": user_id, "event_id": e["event_id"], "event_date": occ_date}, {"_id": 0}
            )
            if mark:
                st = mark["status"]
                per_event[e["event_id"]][st] += 1
            else:
                per_event[e["event_id"]]["unmarked"] += 1
    present = sum(v["present"] for v in per_event.values())
    absent = sum(v["absent"] for v in per_event.values())
    na = sum(v["na"] for v in per_event.values())
    unmarked = sum(v["unmarked"] for v in per_event.values())
    countable = present + absent  # NA and unmarked excluded from %
    pct = round((present / countable) * 100, 1) if countable else 0.0
    return {
        "total_events": total, "present": present, "absent": absent, "na": na, "unmarked": unmarked,
        "attendance_pct": pct, "per_event": list(per_event.values()),
    }


@api.get("/seasons")
async def list_seasons(request: Request):
    await get_current_user(request, db)
    seasons = await db.seasons.find({}, {"_id": 0}).sort("start_date", -1).to_list(200)
    return seasons


@api.post("/seasons")
async def create_season(payload: SeasonIn, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    try:
        s = date.fromisoformat(payload.start_date)
        e = date.fromisoformat(payload.end_date)
        if e < s:
            raise ValueError()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid dates (end must be >= start)")
    doc = payload.model_dump()
    doc.update({
        "season_id": str(uuid.uuid4()),
        "created_by": user["user_id"],
        "created_at": _iso(datetime.now(timezone.utc)),
    })
    await db.seasons.insert_one(doc)
    return _clean(doc)


@api.delete("/seasons/{season_id}")
async def delete_season(season_id: str, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    r = await db.seasons.delete_one({"season_id": season_id})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Season not found")
    return {"ok": True}


@api.get("/seasons/{season_id}/my-report")
async def my_season_report(season_id: str, request: Request):
    user = await get_current_user(request, db)
    season = await db.seasons.find_one({"season_id": season_id}, {"_id": 0})
    if not season:
        raise HTTPException(status_code=404, detail="Season not found")
    believer_only = bool(season.get("is_believer"))
    report = await _compute_report(user["user_id"], season, believer_only=believer_only)
    return {"season": season, "user": {"user_id": user["user_id"], "name": user["name"]}, **report}


@api.get("/seasons/{season_id}/team-report")
async def team_season_report(season_id: str, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin", "team_leader"])
    season = await db.seasons.find_one({"season_id": season_id}, {"_id": 0})
    if not season:
        raise HTTPException(status_code=404, detail="Season not found")
    believer_only = bool(season.get("is_believer"))
    # Scope users
    if user["role"] == "team_leader":
        team = await _my_team_or_403(user)
        users = await db.users.find({"team_id": team["team_id"]}, {"_id": 0, "password_hash": 0}).to_list(500)
    else:
        query = {"is_believer": True} if believer_only else {}
        users = await db.users.find(query, {"_id": 0, "password_hash": 0}).to_list(1000)
    rows = []
    for u in users:
        rep = await _compute_report(u["user_id"], season, believer_only=believer_only)
        rows.append({
            "user_id": u["user_id"], "name": u["name"], "team": u.get("team"),
            "is_believer": u.get("is_believer", False),
            "present": rep["present"], "absent": rep["absent"], "na": rep["na"],
            "unmarked": rep["unmarked"], "total_events": rep["total_events"],
            "attendance_pct": rep["attendance_pct"],
        })
    rows.sort(key=lambda r: r["attendance_pct"], reverse=True)
    return {"season": season, "members": rows}


# --- Believer flag update ---
@api.patch("/admin/users/{uid}/believer")
async def set_believer(uid: str, payload: BelieverUpdate, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    r = await db.users.update_one({"user_id": uid}, {"$set": {"is_believer": payload.is_believer}})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"ok": True}


# --- Tasks (super_admin assigns to members) ---
@api.get("/tasks")
async def list_tasks(request: Request, all_users: bool = False):
    user = await get_current_user(request, db)
    if all_users:
        require_role(user, ["super_admin"])
        tasks = await db.tasks.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    else:
        tasks = await db.tasks.find({"assigned_to": user["user_id"]}, {"_id": 0}).sort("due_date", 1).to_list(500)
    # Enrich with assignee/assigner names
    for t in tasks:
        if t.get("assigned_to"):
            au = await db.users.find_one({"user_id": t["assigned_to"]}, {"_id": 0, "name": 1, "email": 1})
            t["assignee"] = {"name": au["name"], "email": au["email"]} if au else None
    return tasks


@api.post("/tasks")
async def create_task(payload: TaskIn, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    assignee = await db.users.find_one({"user_id": payload.assigned_to}, {"_id": 0})
    if not assignee:
        raise HTTPException(status_code=404, detail="Assignee not found")
    doc = payload.model_dump()
    doc.update({
        "task_id": str(uuid.uuid4()),
        "assigned_by": user["user_id"],
        "status": "pending",
        "completed_at": None,
        "created_at": _iso(datetime.now(timezone.utc)),
    })
    await db.tasks.insert_one(doc)
    return _clean(doc)


@api.patch("/tasks/{task_id}/complete")
async def complete_task(task_id: str, request: Request):
    user = await get_current_user(request, db)
    t = await db.tasks.find_one({"task_id": task_id}, {"_id": 0})
    if not t:
        raise HTTPException(status_code=404, detail="Task not found")
    if t.get("status") == "completed":
        raise HTTPException(status_code=400, detail="Already completed")
    if t.get("assigned_to") != user["user_id"] and user["role"] != "super_admin":
        raise HTTPException(status_code=403, detail="Not your task")
    now = _iso(datetime.now(timezone.utc))
    await db.tasks.update_one({"task_id": task_id}, {"$set": {"status": "completed", "completed_at": now}})
    xp = await award_xp(t["assigned_to"], int(t.get("xp_reward", 0)), f"task:{task_id}")
    updated = await db.tasks.find_one({"task_id": task_id}, {"_id": 0})
    return {"task": updated, "xp": xp}


@api.delete("/tasks/{task_id}")
async def delete_task(task_id: str, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    r = await db.tasks.delete_one({"task_id": task_id})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Task not found")
    return {"ok": True}


# ---------- Rewards / Redemptions ----------
ATTENDANCE_XP_TABLE = [(100, 50), (90, 40), (80, 30), (70, 20), (60, 10), (0, 0)]


def attendance_bonus_xp(pct: float) -> int:
    for threshold, xp in ATTENDANCE_XP_TABLE:
        if pct >= threshold:
            return xp
    return 0


@api.get("/rewards")
async def list_rewards(request: Request):
    await get_current_user(request, db)
    items = await db.rewards.find({"active": True}, {"_id": 0}).sort("cost_xp", 1).to_list(200)
    return items


@api.post("/rewards")
async def create_reward(payload: RewardIn, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    doc = payload.model_dump()
    doc.update({
        "reward_id": str(uuid.uuid4()),
        "created_by": user["user_id"],
        "created_at": _iso(datetime.now(timezone.utc)),
    })
    await db.rewards.insert_one(doc)
    return _clean(doc)


@api.patch("/rewards/{reward_id}")
async def update_reward(reward_id: str, payload: RewardUpdate, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    r = await db.rewards.update_one({"reward_id": reward_id}, {"$set": updates})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="Reward not found")
    return {"ok": True}


@api.delete("/rewards/{reward_id}")
async def delete_reward(reward_id: str, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    r = await db.rewards.delete_one({"reward_id": reward_id})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Reward not found")
    return {"ok": True}


@api.post("/rewards/{reward_id}/redeem")
async def redeem_reward(reward_id: str, request: Request):
    user = await get_current_user(request, db)
    reward = await db.rewards.find_one({"reward_id": reward_id, "active": True}, {"_id": 0})
    if not reward:
        raise HTTPException(status_code=404, detail="Reward not found")
    if reward.get("stock") is not None and reward["stock"] <= 0:
        raise HTTPException(status_code=400, detail="Out of stock")
    if user.get("xp", 0) < reward["cost_xp"]:
        raise HTTPException(status_code=400, detail=f"Not enough XP. Need {reward['cost_xp']} XP.")
    # Deduct XP
    new_xp = user["xp"] - reward["cost_xp"]
    await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"xp": new_xp, "level": level_from_xp(new_xp)}})
    await db.xp_events.insert_one({
        "event_id": str(uuid.uuid4()), "user_id": user["user_id"],
        "amount": -reward["cost_xp"], "reason": f"redeem:{reward_id}",
        "created_at": _iso(datetime.now(timezone.utc)),
    })
    # Decrement stock
    if reward.get("stock") is not None:
        await db.rewards.update_one({"reward_id": reward_id}, {"$inc": {"stock": -1}})
    redemption = {
        "redemption_id": str(uuid.uuid4()),
        "user_id": user["user_id"],
        "user_name": user["name"],
        "reward_id": reward_id,
        "reward_name": reward["name"],
        "cost_xp": reward["cost_xp"],
        "status": "pending",
        "created_at": _iso(datetime.now(timezone.utc)),
        "fulfilled_at": None,
        "fulfilled_by": None,
    }
    await db.redemptions.insert_one(redemption)
    return {"redemption": _clean(redemption), "new_xp": new_xp}


@api.get("/redemptions")
async def list_redemptions(request: Request, all_users: bool = False):
    user = await get_current_user(request, db)
    if all_users:
        require_role(user, ["super_admin"])
        items = await db.redemptions.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    else:
        items = await db.redemptions.find({"user_id": user["user_id"]}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return items


@api.patch("/redemptions/{rid}/fulfill")
async def fulfill_redemption(rid: str, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    r = await db.redemptions.find_one({"redemption_id": rid}, {"_id": 0})
    if not r:
        raise HTTPException(status_code=404, detail="Redemption not found")
    now = _iso(datetime.now(timezone.utc))
    await db.redemptions.update_one(
        {"redemption_id": rid},
        {"$set": {"status": "fulfilled", "fulfilled_at": now, "fulfilled_by": user["user_id"]}},
    )
    return {"ok": True}


# ---------- Team League ----------
async def _team_attendance_pct(team_id: str, days: int = 30) -> dict:
    """Compute team attendance % over the last N days."""
    end = date.today()
    start = end - timedelta(days=days - 1)
    users_in_team = await db.users.find({"team_id": team_id}, {"_id": 0, "user_id": 1}).to_list(1000)
    uids = [u["user_id"] for u in users_in_team]
    if not uids:
        return {"present": 0, "absent": 0, "na": 0, "attendance_pct": 0.0}
    marks = await db.event_attendance.find({
        "user_id": {"$in": uids},
        "event_date": {"$gte": start.isoformat(), "$lte": end.isoformat()},
    }, {"_id": 0}).to_list(50000)
    present = sum(1 for m in marks if m["status"] == "present")
    absent = sum(1 for m in marks if m["status"] == "absent")
    na = sum(1 for m in marks if m["status"] == "na")
    countable = present + absent
    pct = round((present / countable) * 100, 1) if countable else 0.0
    return {"present": present, "absent": absent, "na": na, "attendance_pct": pct}


# ---------- Spartans League (Individual + Team + Season) ----------
async def _current_active_season() -> Optional[dict]:
    today_iso = date.today().isoformat()
    active = await db.seasons.find_one(
        {"start_date": {"$lte": today_iso}, "end_date": {"$gte": today_iso}},
        {"_id": 0},
        sort=[("start_date", -1)],
    )
    if active:
        return active
    # Fallback: most recent season
    return await db.seasons.find_one({}, {"_id": 0}, sort=[("start_date", -1)])


def _season_iso_range(season: dict):
    s = season["start_date"]
    e = season["end_date"]
    return s + "T00:00:00+00:00", e + "T23:59:59+00:00"


async def _user_stats_in_range(user_id: str, start_iso: str, end_iso: str) -> dict:
    """Compute XP, missions, tasks, goals, attendance for a user in date window."""
    xp_agg = await db.xp_events.aggregate([
        {"$match": {"user_id": user_id, "created_at": {"$gte": start_iso, "$lte": end_iso}}},
        {"$group": {"_id": None, "xp": {"$sum": "$amount"}}},
    ]).to_list(1)
    xp = xp_agg[0]["xp"] if xp_agg else 0
    missions = await db.missions.count_documents({
        "user_id": user_id,
        "created_at": {"$gte": start_iso, "$lte": end_iso},
    })
    tasks = await db.tasks.count_documents({
        "assigned_to": user_id, "status": "completed",
        "completed_at": {"$gte": start_iso, "$lte": end_iso},
    })
    goals = await db.goals.count_documents({
        "user_id": user_id, "status": "completed",
        "completed_at": {"$gte": start_iso, "$lte": end_iso},
    })
    return {"xp": xp, "missions": missions, "tasks": tasks, "goals": goals}


async def _user_attendance_in_season(user_id: str, season: dict) -> float:
    marks = await db.event_attendance.find({
        "user_id": user_id,
        "event_date": {"$gte": season["start_date"], "$lte": season["end_date"]},
    }, {"_id": 0}).to_list(50000)
    present = sum(1 for m in marks if m["status"] == "present")
    absent = sum(1 for m in marks if m["status"] == "absent")
    countable = present + absent
    return round((present / countable) * 100, 1) if countable else 0.0


@api.get("/spartans-league/active-season")
async def spartans_active_season(request: Request):
    await get_current_user(request, db)
    s = await _current_active_season()
    return {"season": s}


@api.get("/spartans-league/individual")
async def spartans_individual(request: Request, season_id: Optional[str] = None, limit: int = 100):
    await get_current_user(request, db)
    season = None
    if season_id:
        season = await db.seasons.find_one({"season_id": season_id}, {"_id": 0})
    else:
        season = await _current_active_season()
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(2000)
    rows = []
    if season:
        start_iso, end_iso = _season_iso_range(season)
        for u in users:
            stats = await _user_stats_in_range(u["user_id"], start_iso, end_iso)
            att_pct = await _user_attendance_in_season(u["user_id"], season)
            score = stats["xp"] + stats["missions"] * 5 + stats["tasks"] * 8 + stats["goals"] * 12 + att_pct * 2
            rows.append({
                "user_id": u["user_id"], "name": u["name"],
                "avatar_url": u.get("avatar_url") or u.get("picture"),
                "team": u.get("team"), "level": u.get("level", 1),
                "streak_current": u.get("streak_current", 0),
                "position_badges": u.get("position_badges", []),
                "club_type": u.get("club_type"),
                "nexus_id": u.get("nexus_id"),
                "xp": stats["xp"], "missions": stats["missions"], "tasks": stats["tasks"],
                "goals": stats["goals"], "attendance_pct": att_pct,
                "score": round(score, 1),
            })
    else:
        # No season: fall back to all-time XP
        for u in users:
            rows.append({
                "user_id": u["user_id"], "name": u["name"],
                "avatar_url": u.get("avatar_url") or u.get("picture"),
                "team": u.get("team"), "level": u.get("level", 1),
                "streak_current": u.get("streak_current", 0),
                "position_badges": u.get("position_badges", []),
                "club_type": u.get("club_type"),
                "nexus_id": u.get("nexus_id"),
                "xp": u.get("xp", 0), "missions": 0, "tasks": 0, "goals": 0,
                "attendance_pct": 0.0, "score": u.get("xp", 0),
            })
    rows.sort(key=lambda r: r["score"], reverse=True)
    rows = rows[:limit]
    for i, r in enumerate(rows):
        r["rank"] = i + 1
    return {"season": season, "rows": rows}


@api.get("/spartans-league/team")
async def spartans_team_league(request: Request, season_id: Optional[str] = None):
    """
    Team League formula:
      40% XP + 25% Attendance + 15% Missions + 10% Tasks + 10% Goals
    All metrics scoped to the active/specified season.
    """
    await get_current_user(request, db)
    season = None
    if season_id:
        season = await db.seasons.find_one({"season_id": season_id}, {"_id": 0})
    else:
        season = await _current_active_season()

    teams = await db.teams.find({}, {"_id": 0}).to_list(200)
    raw = []
    if season:
        start_iso, end_iso = _season_iso_range(season)
    for t in teams:
        members = await db.users.find({"team_id": t["team_id"]}, {"_id": 0, "password_hash": 0}).to_list(500)
        if not members:
            continue
        mids = [m["user_id"] for m in members]
        # xp/missions/tasks/goals aggregate over members
        if season:
            xp_agg = await db.xp_events.aggregate([
                {"$match": {"user_id": {"$in": mids}, "created_at": {"$gte": start_iso, "$lte": end_iso}}},
                {"$group": {"_id": None, "xp": {"$sum": "$amount"}}},
            ]).to_list(1)
            xp = xp_agg[0]["xp"] if xp_agg else 0
            missions = await db.missions.count_documents({
                "user_id": {"$in": mids}, "created_at": {"$gte": start_iso, "$lte": end_iso},
            })
            tasks = await db.tasks.count_documents({
                "assigned_to": {"$in": mids}, "status": "completed",
                "completed_at": {"$gte": start_iso, "$lte": end_iso},
            })
            goals = await db.goals.count_documents({
                "user_id": {"$in": mids}, "status": "completed",
                "completed_at": {"$gte": start_iso, "$lte": end_iso},
            })
            att = await _team_attendance_pct_in_range(t["team_id"], season["start_date"], season["end_date"])
        else:
            xp = sum(m.get("xp", 0) for m in members)
            missions = await db.missions.count_documents({"user_id": {"$in": mids}})
            tasks = await db.tasks.count_documents({"assigned_to": {"$in": mids}, "status": "completed"})
            goals = await db.goals.count_documents({"user_id": {"$in": mids}, "status": "completed"})
            att = 0.0
        raw.append({
            "team_id": t["team_id"], "name": t["name"], "members": len(members),
            "xp": xp, "missions": missions, "tasks": tasks, "goals": goals,
            "attendance_pct": att, "leader_id": t.get("leader_id"),
            "leader_name": None, "leader_avatar_url": None, "leader_badges": [],
            "mission_pct": min(100, round((missions / max(1, len(members) * 20)) * 100, 1)),
        })
    # Attach leader info
    for r in raw:
        if r.get("leader_id"):
            lu = await db.users.find_one({"user_id": r["leader_id"]}, {"_id": 0, "name": 1, "avatar_url": 1, "picture": 1, "position_badges": 1})
            if lu:
                r["leader_name"] = lu.get("name")
                r["leader_avatar_url"] = lu.get("avatar_url") or lu.get("picture")
                r["leader_badges"] = lu.get("position_badges", [])
    # Normalize each metric to 0..100 relative to max
    max_xp = max((r["xp"] for r in raw), default=0) or 1
    max_mis = max((r["missions"] for r in raw), default=0) or 1
    max_tsk = max((r["tasks"] for r in raw), default=0) or 1
    max_gol = max((r["goals"] for r in raw), default=0) or 1
    for r in raw:
        score = (
            (r["xp"] / max_xp) * 40 +
            (r["attendance_pct"] / 100.0) * 25 +
            (r["missions"] / max_mis) * 15 +
            (r["tasks"] / max_tsk) * 10 +
            (r["goals"] / max_gol) * 10
        )
        r["score"] = round(score, 2)
        r["breakdown"] = {
            "xp_pts": round((r["xp"] / max_xp) * 40, 2),
            "attendance_pts": round((r["attendance_pct"] / 100.0) * 25, 2),
            "missions_pts": round((r["missions"] / max_mis) * 15, 2),
            "tasks_pts": round((r["tasks"] / max_tsk) * 10, 2),
            "goals_pts": round((r["goals"] / max_gol) * 10, 2),
        }
    raw.sort(key=lambda r: r["score"], reverse=True)
    for i, r in enumerate(raw):
        r["rank"] = i + 1
    return {
        "season": season, "teams": raw,
        "weights": {"xp": 40, "attendance": 25, "missions": 15, "tasks": 10, "goals": 10},
    }


async def _team_attendance_pct_in_range(team_id: str, start_date: str, end_date: str) -> float:
    team_users = await db.users.find({"team_id": team_id}, {"_id": 0, "user_id": 1}).to_list(1000)
    uids = [t["user_id"] for t in team_users]
    if not uids:
        return 0.0
    marks = await db.event_attendance.find({
        "user_id": {"$in": uids},
        "event_date": {"$gte": start_date, "$lte": end_date},
    }, {"_id": 0}).to_list(50000)
    present = sum(1 for m in marks if m["status"] == "present")
    absent = sum(1 for m in marks if m["status"] == "absent")
    countable = present + absent
    return round((present / countable) * 100, 1) if countable else 0.0


@api.get("/team-league")
async def team_league(request: Request):
    await get_current_user(request, db)
    teams = await db.teams.find({}, {"_id": 0}).to_list(200)
    rows = []
    week_start_iso = (date.today() - timedelta(days=7)).isoformat()
    month_start_iso = (date.today() - timedelta(days=30)).isoformat()
    for t in teams:
        members = await db.users.find({"team_id": t["team_id"]}, {"_id": 0, "password_hash": 0}).to_list(500)
        if not members:
            continue
        xp = sum(m.get("xp", 0) for m in members)
        streak = max((m.get("streak_current", 0) for m in members), default=0)
        # Weekly XP
        mids = [m["user_id"] for m in members]
        weekly_xp_agg = await db.xp_events.aggregate([
            {"$match": {"user_id": {"$in": mids}, "created_at": {"$gte": (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()}}},
            {"$group": {"_id": None, "xp": {"$sum": "$amount"}}},
        ]).to_list(1)
        weekly_xp = weekly_xp_agg[0]["xp"] if weekly_xp_agg else 0
        monthly_xp_agg = await db.xp_events.aggregate([
            {"$match": {"user_id": {"$in": mids}, "created_at": {"$gte": (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()}}},
            {"$group": {"_id": None, "xp": {"$sum": "$amount"}}},
        ]).to_list(1)
        monthly_xp = monthly_xp_agg[0]["xp"] if monthly_xp_agg else 0
        # Weekly + Monthly attendance
        weekly_att = await _team_attendance_pct(t["team_id"], days=7)
        monthly_att = await _team_attendance_pct(t["team_id"], days=30)
        bonus = attendance_bonus_xp(monthly_att["attendance_pct"])
        rows.append({
            "team_id": t["team_id"], "name": t["name"], "members": len(members),
            "xp": xp, "streak": streak,
            "weekly_xp": weekly_xp, "monthly_xp": monthly_xp,
            "weekly_attendance_pct": weekly_att["attendance_pct"],
            "monthly_attendance_pct": monthly_att["attendance_pct"],
            "attendance_bonus_xp": bonus,
            "leader": t.get("leader_id"),
        })
    rows.sort(key=lambda r: (r["monthly_attendance_pct"] + r["xp"] / 1000), reverse=True)
    for i, r in enumerate(rows):
        r["rank"] = i + 1
    return {
        "teams": rows,
        "attendance_xp_table": [{"threshold_pct": t, "xp_reward": x} for t, x in ATTENDANCE_XP_TABLE],
    }


# ---------- Profile Completion ----------
PROFILE_FIELDS = ["name", "email", "avatar_url", "team_id", "phone", "bio",
                  "dob", "gender", "marital_status", "city", "state",
                  "club_type",
                  "favourite_food", "favourite_place", "favourite_hobby"]
PROFILE_COMPLETION_XP = 50


def _profile_completion(user: dict):
    filled = sum(1 for f in PROFILE_FIELDS if user.get(f))
    total = len(PROFILE_FIELDS)
    pct = round((filled / total) * 100, 0)
    missing = [f for f in PROFILE_FIELDS if not user.get(f)]
    return {"filled": filled, "total": total, "pct": pct, "missing": missing,
            "completion_xp_awarded": bool(user.get("profile_completed_awarded"))}


@api.get("/profile/completion")
async def profile_completion(request: Request):
    user = await get_current_user(request, db)
    return _profile_completion(user)


@api.patch("/profile")
async def update_profile(payload: ProfileUpdate, request: Request):
    user = await get_current_user(request, db)
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    await db.users.update_one({"user_id": user["user_id"]}, {"$set": updates})
    fresh = await db.users.find_one({"user_id": user["user_id"]}, {"_id": 0, "password_hash": 0})
    comp = _profile_completion(fresh)
    xp_awarded = None
    if comp["pct"] >= 100 and not fresh.get("profile_completed_awarded"):
        await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"profile_completed_awarded": True}})
        xp_awarded = await award_xp(user["user_id"], PROFILE_COMPLETION_XP, "profile_completion")
    return {"user": fresh, "completion": _profile_completion(fresh), "xp": xp_awarded}


@api.get("/teams/public")
async def list_teams_public(request: Request):
    """Any user can see the list of active teams for self-selection."""
    await get_current_user(request, db)
    teams = await db.teams.find({}, {"_id": 0, "team_id": 1, "name": 1, "leader_id": 1}).sort("name", 1).to_list(200)
    return teams


@api.post("/profile/join-team")
async def self_join_team(request: Request, team_id: str):
    """Users self-select a team during profile completion."""
    user = await get_current_user(request, db)
    team = await db.teams.find_one({"team_id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    await db.users.update_one(
        {"user_id": user["user_id"]},
        {"$set": {"team_id": team_id, "team": team["name"]}},
    )
    return {"ok": True, "team": team}


# ---------- Exports (CSV + PDF + XLSX) ----------
import io
import csv


def _sanitize_cell(v):
    """Neutralize CSV formula injection: prefix a leading = + - @ tab with a single quote."""
    if isinstance(v, str) and v and v[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + v
    return v


def _csv_response(filename: str, headers: list, rows: list):
    from fastapi.responses import StreamingResponse
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    for r in rows:
        w.writerow([_sanitize_cell(c) for c in r])
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _xlsx_response(filename: str, title: str, headers: list, rows: list):
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Report")[:31]
    header_font = Font(bold=True, color="FFFFFFFF", size=11)
    header_fill = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=str(h))
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=_sanitize_cell(val))
    # Auto width
    for col in ws.columns:
        length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(60, max(10, length + 2))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _export_response(fmt: str, filename_stem: str, title: str, headers: list, rows: list):
    """Unified dispatch: fmt in {csv, pdf, xlsx, excel}."""
    fmt = (fmt or "csv").lower()
    if fmt == "pdf":
        return _pdf_response(filename_stem + ".pdf", title, headers, rows)
    if fmt in ("xlsx", "excel"):
        return _xlsx_response(filename_stem + ".xlsx", title, headers, rows)
    return _csv_response(filename_stem + ".csv", headers, rows)


def _pdf_response(filename: str, title: str, headers: list, rows: list):
    from fastapi.responses import StreamingResponse
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title=title)
    styles = getSampleStyleSheet()
    elems = [Paragraph(f"<b>SPARTANS GROWTH LEAGUE</b>", styles["Title"]),
             Paragraph(title, styles["Heading2"]),
             Paragraph(f"Generated: {datetime.now(LOCAL_TZ).strftime('%Y-%m-%d %H:%M IST')}", styles["Normal"]),
             Spacer(1, 10)]
    data = [headers] + [[str(c) if c is not None else "" for c in r] for r in rows]
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EAB308")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F9FAFB"), colors.white]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
    ]))
    elems.append(tbl)
    doc.build(elems)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api.get("/exports/team-performance")
async def export_team_performance(request: Request, format: str = "csv"):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin", "team_leader"])
    league = await team_league(request)
    headers = ["Rank", "Team", "Members", "Total XP", "Weekly XP", "Monthly XP",
               "Weekly Attn %", "Monthly Attn %", "Bonus XP", "Best Streak"]
    rows = [[r["rank"], r["name"], r["members"], r["xp"], r["weekly_xp"], r["monthly_xp"],
             r["weekly_attendance_pct"], r["monthly_attendance_pct"], r["attendance_bonus_xp"],
             r["streak"]] for r in league["teams"]]
    filename = f"team-performance-{date.today().isoformat()}"
    return _export_response(format, filename, "Team Performance Report", headers, rows)


@api.get("/exports/attendance")
async def export_attendance(request: Request, format: str = "csv", season_id: Optional[str] = None):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin", "team_leader"])
    if not season_id:
        # Latest season if any
        s = await db.seasons.find_one({}, {"_id": 0}, sort=[("start_date", -1)])
        if not s:
            raise HTTPException(status_code=400, detail="No season available. Create one or pass season_id.")
        season_id = s["season_id"]
    # Reuse team-report logic
    season = await db.seasons.find_one({"season_id": season_id}, {"_id": 0})
    if not season:
        raise HTTPException(status_code=404, detail="Season not found")
    believer_only = bool(season.get("is_believer"))
    if user["role"] == "team_leader":
        team = await _my_team_or_403(user)
        users = await db.users.find({"team_id": team["team_id"]}, {"_id": 0, "password_hash": 0}).to_list(1000)
    else:
        q = {"is_believer": True} if believer_only else {}
        users = await db.users.find(q, {"_id": 0, "password_hash": 0}).to_list(2000)
    rows = []
    for u in users:
        rep = await _compute_report(u["user_id"], season, believer_only=believer_only)
        rows.append([u["name"], u.get("team") or "-", rep["present"], rep["absent"], rep["na"],
                     rep["unmarked"], rep["total_events"], rep["attendance_pct"]])
    rows.sort(key=lambda r: r[7], reverse=True)
    headers = ["Name", "Team", "Present", "Absent", "N/A", "Unmarked", "Total Events", "Attendance %"]
    filename = f"attendance-{season['name'].replace(' ', '_')}-{date.today().isoformat()}"
    return _export_response(format, filename, f"Attendance Report — {season['name']}", headers, rows)


@api.get("/exports/xp-leaderboard")
async def export_xp_leaderboard(request: Request, format: str = "csv", scope: str = "all"):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin", "team_leader"])
    lb = await leaderboard(request, scope=scope, limit=500)
    headers = ["Rank", "Name", "Team", "Level", "XP", "Streak"]
    rows = [[r["rank"], r["name"], r.get("team") or "-", r["level"], r["xp"], r["streak_current"]]
            for r in lb]
    filename = f"xp-leaderboard-{scope}-{date.today().isoformat()}"
    return _export_response(format, filename, f"XP Leaderboard — {scope.title()}", headers, rows)


@api.get("/exports/daily")
async def export_daily(request: Request, format: str = "csv", day: Optional[str] = None):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin", "team_leader"])
    d = day or date.today().isoformat()
    # XP events per user for that day
    day_start = _iso(datetime.combine(date.fromisoformat(d), datetime.min.time()).replace(tzinfo=timezone.utc))
    day_end = _iso(datetime.combine(date.fromisoformat(d) + timedelta(days=1), datetime.min.time()).replace(tzinfo=timezone.utc))
    pipeline = [
        {"$match": {"created_at": {"$gte": day_start, "$lt": day_end}}},
        {"$group": {"_id": "$user_id", "xp": {"$sum": "$amount"}, "events": {"$sum": 1}}},
    ]
    events = await db.xp_events.aggregate(pipeline).to_list(2000)
    uids = [e["_id"] for e in events]
    users = {u["user_id"]: u for u in await db.users.find({"user_id": {"$in": uids}}, {"_id": 0, "password_hash": 0}).to_list(2000)}
    rows = []
    for e in events:
        u = users.get(e["_id"])
        if not u:
            continue
        rows.append([u["name"], u.get("team") or "-", e["events"], e["xp"]])
    rows.sort(key=lambda r: r[3], reverse=True)
    headers = ["Name", "Team", "Actions", "XP Earned"]
    filename = f"daily-report-{d}"
    return _export_response(format, filename, f"Daily Report — {d}", headers, rows)


# ---------- Unified Report Exports (Missions/Tasks/Goals/Followups/League) ----------
async def _scoped_user_ids(user: dict) -> Optional[List[str]]:
    """Return None for global scope, otherwise a list of member ids for team_leader."""
    if user["role"] == "team_leader":
        team = await _my_team_or_403(user)
        return [m["user_id"] async for m in db.users.find({"team_id": team["team_id"]}, {"_id": 0, "user_id": 1})]
    return None


async def _uid_to_name(uids: List[str]) -> dict:
    us = await db.users.find({"user_id": {"$in": uids}}, {"_id": 0, "user_id": 1, "name": 1, "team": 1}).to_list(len(uids) or 1)
    return {u["user_id"]: u for u in us}


@api.get("/exports/missions")
async def export_missions(request: Request, format: str = "csv"):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin", "team_leader"])
    scoped = await _scoped_user_ids(user)
    q = {"user_id": {"$in": scoped}} if scoped is not None else {}
    missions = await db.missions.find(q, {"_id": 0}).sort("created_at", -1).to_list(5000)
    uids = list({m["user_id"] for m in missions})
    umap = await _uid_to_name(uids)
    headers = ["Date", "Spartan", "Team", "Prospect", "Mobile", "Status", "GPS"]
    rows = []
    for m in missions:
        u = umap.get(m["user_id"], {})
        rows.append([
            m.get("created_at", "")[:10],
            u.get("name", "?"),
            u.get("team", "-"),
            m.get("prospect_name", ""),
            m.get("mobile_number", ""),
            m.get("status", ""),
            f"{m.get('lat', '')},{m.get('lng', '')}" if m.get("lat") else "-",
        ])
    filename = f"missions-{date.today().isoformat()}"
    return _export_response(format, filename, "Missions Report", headers, rows)


@api.get("/exports/tasks")
async def export_tasks(request: Request, format: str = "csv", status: Optional[str] = None):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin", "team_leader"])
    scoped = await _scoped_user_ids(user)
    q = {}
    if scoped is not None:
        q["assigned_to"] = {"$in": scoped}
    if status:
        q["status"] = status
    tasks = await db.tasks.find(q, {"_id": 0}).sort("due_date", 1).to_list(5000)
    uids = list({t["assigned_to"] for t in tasks})
    umap = await _uid_to_name(uids)
    headers = ["Title", "Assignee", "Team", "Due Date", "Status", "XP Reward", "Completed At"]
    rows = []
    for t in tasks:
        u = umap.get(t["assigned_to"], {})
        rows.append([t.get("title", ""), u.get("name", "?"), u.get("team", "-"),
                     t.get("due_date", ""), t.get("status", ""), t.get("xp_reward", 0),
                     (t.get("completed_at") or "")[:19]])
    filename = f"tasks-{date.today().isoformat()}"
    return _export_response(format, filename, "Tasks Report", headers, rows)


@api.get("/exports/goals")
async def export_goals(request: Request, format: str = "csv"):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin", "team_leader"])
    scoped = await _scoped_user_ids(user)
    q = {"user_id": {"$in": scoped}} if scoped is not None else {}
    goals = await db.goals.find(q, {"_id": 0}).sort("created_at", -1).to_list(5000)
    uids = list({g["user_id"] for g in goals})
    umap = await _uid_to_name(uids)
    headers = ["Title", "Spartan", "Team", "Period", "Progress", "Target", "Status", "XP Reward", "Completed At"]
    rows = []
    for g in goals:
        u = umap.get(g["user_id"], {})
        rows.append([g.get("title", ""), u.get("name", "?"), u.get("team", "-"),
                     g.get("period", ""), g.get("progress", 0), g.get("target", 0),
                     g.get("status", ""), g.get("xp_reward", 0),
                     (g.get("completed_at") or "")[:19]])
    filename = f"goals-{date.today().isoformat()}"
    return _export_response(format, filename, "Goals Report", headers, rows)


@api.get("/exports/followups")
async def export_followups(request: Request, format: str = "csv"):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin", "team_leader"])
    scoped = await _scoped_user_ids(user)
    q = {"user_id": {"$in": scoped}} if scoped is not None else {}
    fus = await db.followups.find(q, {"_id": 0}).sort("due_date", 1).to_list(5000)
    uids = list({f["user_id"] for f in fus})
    umap = await _uid_to_name(uids)
    headers = ["Title", "Spartan", "Team", "Due Date", "Status", "Notes"]
    rows = []
    for f in fus:
        u = umap.get(f["user_id"], {})
        rows.append([f.get("title", ""), u.get("name", "?"), u.get("team", "-"),
                     f.get("due_date", ""), f.get("status", ""), (f.get("notes") or "")[:100]])
    filename = f"followups-{date.today().isoformat()}"
    return _export_response(format, filename, "Follow-Ups Report", headers, rows)


@api.get("/exports/spartans-league")
async def export_spartans_league(request: Request, format: str = "csv", scope: str = "individual", season_id: Optional[str] = None):
    """scope: individual | team"""
    user = await get_current_user(request, db)
    require_role(user, ["super_admin", "team_leader"])
    if scope == "team":
        data = await spartans_team_league(request, season_id=season_id)
        headers = ["Rank", "Team", "Members", "Score", "XP", "Attendance %", "Missions", "Tasks", "Goals"]
        rows = [[r["rank"], r["name"], r["members"], r["score"], r["xp"],
                 r["attendance_pct"], r["missions"], r["tasks"], r["goals"]] for r in data["teams"]]
        title = "Spartans League — Team"
        filename = f"spartans-league-team-{date.today().isoformat()}"
    else:
        data = await spartans_individual(request, season_id=season_id, limit=500)
        headers = ["Rank", "Name", "Team", "Level", "Score", "XP", "Missions", "Tasks", "Goals", "Attendance %"]
        rows = [[r["rank"], r["name"], r.get("team") or "-", r["level"], r["score"],
                 r["xp"], r["missions"], r["tasks"], r["goals"], r["attendance_pct"]] for r in data["rows"]]
        title = "Spartans League — Individual"
        filename = f"spartans-league-individual-{date.today().isoformat()}"
    return _export_response(format, filename, title, headers, rows)


# ---------- Celebrations ----------
def _md(iso_date: str) -> str:
    """Return MM-DD from a YYYY-MM-DD string."""
    try:
        return iso_date[5:10]
    except Exception:
        return ""


@api.get("/celebrations/me")
async def my_celebrations(request: Request):
    user = await get_current_user(request, db)
    today_md = date.today().strftime("%m-%d")
    is_birthday = user.get("dob") and _md(user["dob"]) == today_md
    is_anniversary = (
        user.get("marital_status") == "married"
        and user.get("anniversary_date")
        and _md(user["anniversary_date"]) == today_md
    )
    return {"is_birthday": bool(is_birthday), "is_anniversary": bool(is_anniversary),
            "user": {"user_id": user["user_id"], "name": user["name"]}}


@api.get("/celebrations/today")
async def team_celebrations(request: Request):
    """List all users whose birthday or anniversary is today."""
    await get_current_user(request, db)
    today_md = date.today().strftime("%m-%d")
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(2000)
    birthdays = []
    anniversaries = []
    for u in users:
        if u.get("dob") and _md(u["dob"]) == today_md:
            birthdays.append({"user_id": u["user_id"], "name": u["name"],
                              "avatar_url": u.get("avatar_url") or u.get("avatar_photo"),
                              "team": u.get("team")})
        if (u.get("marital_status") == "married"
                and u.get("anniversary_date")
                and _md(u["anniversary_date"]) == today_md):
            anniversaries.append({"user_id": u["user_id"], "name": u["name"],
                                  "avatar_url": u.get("avatar_url") or u.get("avatar_photo"),
                                  "anniversary_photo": u.get("anniversary_photo"),
                                  "team": u.get("team")})
    return {"date": date.today().isoformat(), "birthdays": birthdays, "anniversaries": anniversaries}


# ---------- Position Badges (admin-controlled, publicly visible) ----------
@api.get("/position-badges/catalog")
async def badge_catalog(request: Request):
    await get_current_user(request, db)
    return {"badges": POSITION_BADGES}


@api.patch("/admin/users/{uid}/position-badges")
async def set_position_badges(uid: str, payload: PositionBadgesUpdate, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    r = await db.users.update_one({"user_id": uid}, {"$set": {"position_badges": payload.badges}})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"ok": True, "position_badges": payload.badges}


# ---------- Team Leader: Add Member to own team ----------
@api.post("/team-leader/add-member")
async def leader_add_member(payload: AddMemberIn, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["team_leader", "super_admin"])
    team = None
    if user["role"] == "team_leader":
        team = await db.teams.find_one({"leader_id": user["user_id"]}, {"_id": 0})
        if not team:
            raise HTTPException(status_code=404, detail="You do not lead a team")
    email = payload.email.lower().strip()
    exists = await db.users.find_one({"email": email})
    if exists:
        raise HTTPException(status_code=400, detail="Email already registered")
    new_uid = f"user_{uuid.uuid4().hex[:12]}"
    doc = {
        "user_id": new_uid, "email": email, "name": payload.name.strip(),
        "password_hash": hash_password(payload.password), "role": "member",
        "avatar_url": None, "picture": None, "xp": 0, "level": 1,
        "streak_current": 0, "streak_longest": 0, "last_checkin_date": None,
        "team": team["name"] if team else None,
        "team_id": team["team_id"] if team else None,
        "phone": payload.phone, "badges": [], "position_badges": [],
        "created_at": _iso(datetime.now(timezone.utc)), "active": True,
        "added_by_leader": user["user_id"],
    }
    await db.users.insert_one(doc)
    # Auto-assign active admin goal templates
    try:
        for tpl in await db.goal_templates.find({"active": True}, {"_id": 0}).to_list(200):
            await _sync_template_to_users(tpl)
    except Exception as e:
        logger.warning(f"Failed to auto-assign goal templates: {e}")
    return {"ok": True, "user": _clean(doc)}


# ---------- Object Storage (Profile Photo Upload) ----------
import requests as _requests

STORAGE_URL = "https://integrations.emergentagent.com/objstore/api/v1/storage"
STORAGE_APP_NAME = "spartans-growth-league"
_storage_state = {"key": None}


def _init_storage():
    if _storage_state["key"]:
        return _storage_state["key"]
    key = os.environ.get("EMERGENT_LLM_KEY")
    if not key:
        raise HTTPException(status_code=500, detail="Object storage not configured (EMERGENT_LLM_KEY missing)")
    resp = _requests.post(f"{STORAGE_URL}/init", json={"emergent_key": key}, timeout=30)
    resp.raise_for_status()
    _storage_state["key"] = resp.json()["storage_key"]
    return _storage_state["key"]


def _put_object(path: str, data: bytes, content_type: str) -> dict:
    key = _init_storage()
    r = _requests.put(f"{STORAGE_URL}/objects/{path}",
                      headers={"X-Storage-Key": key, "Content-Type": content_type},
                      data=data, timeout=120)
    if r.status_code == 403:
        _storage_state["key"] = None
        key = _init_storage()
        r = _requests.put(f"{STORAGE_URL}/objects/{path}",
                          headers={"X-Storage-Key": key, "Content-Type": content_type},
                          data=data, timeout=120)
    r.raise_for_status()
    return r.json()


def _get_object(path: str):
    key = _init_storage()
    r = _requests.get(f"{STORAGE_URL}/objects/{path}",
                      headers={"X-Storage-Key": key}, timeout=60)
    if r.status_code == 403:
        _storage_state["key"] = None
        key = _init_storage()
        r = _requests.get(f"{STORAGE_URL}/objects/{path}",
                          headers={"X-Storage-Key": key}, timeout=60)
    r.raise_for_status()
    return r.content, r.headers.get("Content-Type", "application/octet-stream")


IMAGE_MIME = {"image/jpeg", "image/png", "image/webp", "image/gif"}
IMAGE_EXT = {"image/jpeg": "jpg", "image/png": "png", "image/webp": "webp", "image/gif": "gif"}
MAX_IMAGE_BYTES = 2 * 1024 * 1024  # 2 MB


@api.post("/uploads/avatar")
async def upload_avatar(request: Request, file: UploadFile = File(...)):
    user = await get_current_user(request, db)
    ctype = (file.content_type or "").lower()
    if ctype not in IMAGE_MIME:
        raise HTTPException(status_code=400, detail="Only image files (jpg, png, webp, gif) allowed")
    data = await file.read()
    if len(data) > MAX_IMAGE_BYTES:
        raise HTTPException(status_code=400, detail="Image must be under 2 MB")
    ext = IMAGE_EXT[ctype]
    path = f"{STORAGE_APP_NAME}/avatars/{user['user_id']}/{uuid.uuid4()}.{ext}"
    result = _put_object(path, data, ctype)
    stored = result.get("path", path)
    file_id = str(uuid.uuid4())
    await db.files.insert_one({
        "file_id": file_id, "storage_path": stored, "user_id": user["user_id"],
        "purpose": "avatar", "content_type": ctype, "size": len(data),
        "is_deleted": False, "created_at": _iso(datetime.now(timezone.utc)),
    })
    # Update user profile
    avatar_url = f"/api/files/{file_id}"
    await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"avatar_url": avatar_url}})
    return {"file_id": file_id, "url": avatar_url, "storage_path": stored, "size": len(data)}


@api.get("/files/{file_id}")
async def download_file(file_id: str, request: Request, auth: Optional[str] = None):
    # Support ?auth=<token> because <img> tags cannot pass Authorization headers.
    caller_user_id = None
    caller_role = None
    if auth:
        try:
            claims = jwt.decode(auth, get_jwt_secret(), algorithms=["HS256"])
            if claims.get("type") != "access":
                raise HTTPException(status_code=401, detail="Invalid token type")
            caller_user_id = claims["sub"]
            u = await db.users.find_one({"user_id": caller_user_id}, {"_id": 0, "role": 1})
            caller_role = u.get("role") if u else None
        except HTTPException:
            raise
        except Exception:
            raise HTTPException(status_code=401, detail="Invalid token")
    else:
        current = await get_current_user(request, db)
        caller_user_id = current["user_id"]
        caller_role = current["role"]
    rec = await db.files.find_one({"file_id": file_id, "is_deleted": False}, {"_id": 0})
    if not rec:
        raise HTTPException(status_code=404, detail="File not found")
    # ACL: super_admin can view any avatar. Otherwise, avatars are public within the app
    # (all authenticated users) because they render throughout leaderboards & rosters.
    # Non-avatar file purposes (future) get strict owner-only access.
    if rec.get("purpose") != "avatar" and caller_role != "super_admin" and rec.get("user_id") != caller_user_id:
        raise HTTPException(status_code=403, detail="Forbidden")
    data, ctype = _get_object(rec["storage_path"])
    return Response(content=data, media_type=rec.get("content_type") or ctype)


# ---------- Admin User Management ----------
class AdminUserUpdate(BaseModel):
    name: Optional[str] = None
    team_id: Optional[str] = None  # empty string = unassign
    role: Optional[Literal["member", "team_leader", "super_admin"]] = None


@api.patch("/admin/users/{uid}")
async def admin_edit_user(uid: str, payload: AdminUserUpdate, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    target = await db.users.find_one({"user_id": uid}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    updates = {}
    if payload.name is not None:
        updates["name"] = payload.name.strip()
    if payload.team_id is not None:
        if payload.team_id == "":
            updates["team_id"] = None
            updates["team"] = None
        else:
            team = await db.teams.find_one({"team_id": payload.team_id}, {"_id": 0})
            if not team:
                raise HTTPException(status_code=404, detail="Team not found")
            updates["team_id"] = team["team_id"]
            updates["team"] = team["name"]
    if payload.role is not None:
        updates["role"] = payload.role
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    await db.users.update_one({"user_id": uid}, {"$set": updates})
    # If demoting away from team_leader, unassign leadership
    if payload.role and payload.role != "team_leader":
        await db.teams.update_many({"leader_id": uid}, {"$set": {"leader_id": None}})
    return {"ok": True}


@api.delete("/admin/users/{uid}")
async def admin_delete_user(uid: str, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    if uid == user["user_id"]:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    target = await db.users.find_one({"user_id": uid}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    if target.get("role") == "super_admin":
        raise HTTPException(status_code=400, detail="Cannot delete another super admin")
    # Cascade: remove records owned by this user so exports/leagues don't show orphans.
    await db.users.delete_one({"user_id": uid})
    await db.teams.update_many({"leader_id": uid}, {"$set": {"leader_id": None}})
    for coll, key in [
        ("prospects", "user_id"), ("followups", "user_id"), ("missions", "user_id"),
        ("goals", "user_id"), ("xp_events", "user_id"), ("event_attendance", "user_id"),
        ("attendance", "user_id"), ("redemptions", "user_id"), ("user_business", "user_id"),
        ("tasks", "assigned_to"), ("files", "user_id"),
    ]:
        await db[coll].delete_many({key: uid})
    return {"ok": True}


class TeamIn(BaseModel):
    name: str = Field(min_length=1, max_length=60)
    leader_id: Optional[str] = None


class TeamUpdate(BaseModel):
    name: Optional[str] = None
    leader_id: Optional[str] = None


@api.post("/admin/teams")
async def admin_create_team(payload: TeamIn, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    existing = await db.teams.find_one({"name": payload.name.strip()})
    if existing:
        raise HTTPException(status_code=400, detail="A team with that name already exists")
    tid = str(uuid.uuid4())
    doc = {"team_id": tid, "name": payload.name.strip(),
           "leader_id": payload.leader_id, "created_at": _iso(datetime.now(timezone.utc))}
    await db.teams.insert_one(doc)
    if payload.leader_id:
        await db.users.update_one({"user_id": payload.leader_id},
                                  {"$set": {"role": "team_leader", "team_id": tid, "team": doc["name"]}})
    return _clean(doc)


@api.patch("/admin/teams/{team_id}")
async def admin_update_team(team_id: str, payload: TeamUpdate, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    team = await db.teams.find_one({"team_id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    updates = {}
    if payload.name is not None:
        updates["name"] = payload.name.strip()
    if payload.leader_id is not None:
        # Set new leader; ensure they exist
        if payload.leader_id == "":
            updates["leader_id"] = None
        else:
            u = await db.users.find_one({"user_id": payload.leader_id}, {"_id": 0})
            if not u:
                raise HTTPException(status_code=404, detail="Leader user not found")
            # Remove leader flag from previous leader
            if team.get("leader_id") and team["leader_id"] != payload.leader_id:
                await db.users.update_one({"user_id": team["leader_id"]},
                                          {"$set": {"role": "member"}})
            updates["leader_id"] = payload.leader_id
            await db.users.update_one({"user_id": payload.leader_id},
                                      {"$set": {"role": "team_leader", "team_id": team_id,
                                                "team": updates.get("name") or team["name"]}})
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    await db.teams.update_one({"team_id": team_id}, {"$set": updates})
    # Cascade name change to member records
    if "name" in updates:
        await db.users.update_many({"team_id": team_id}, {"$set": {"team": updates["name"]}})
    return {"ok": True}


@api.delete("/admin/teams/{team_id}")
async def admin_delete_team(team_id: str, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    team = await db.teams.find_one({"team_id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    n_members = await db.users.count_documents({"team_id": team_id})
    if n_members > 0:
        raise HTTPException(status_code=400, detail=f"Team has {n_members} members; move them first")
    await db.teams.delete_one({"team_id": team_id})
    return {"ok": True}


# ---------- Business Volume (PV) & Earnings ----------
@api.patch("/admin/seasons/{season_id}")
async def admin_update_season(season_id: str, payload: SeasonUpdate, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    updates = {k: v for k, v in payload.model_dump(exclude_unset=True).items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    r = await db.seasons.update_one({"season_id": season_id}, {"$set": updates})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="Season not found")
    return {"ok": True}


@api.post("/admin/users/{uid}/business")
async def admin_set_user_business(uid: str, payload: UserBusinessUpdate, request: Request):
    """Set a user's PV & Earnings for a specific season."""
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    if not await db.users.find_one({"user_id": uid}, {"_id": 0, "user_id": 1}):
        raise HTTPException(status_code=404, detail="User not found")
    if not await db.seasons.find_one({"season_id": payload.season_id}, {"_id": 0, "season_id": 1}):
        raise HTTPException(status_code=404, detail="Season not found")
    key = {"user_id": uid, "season_id": payload.season_id}
    now = _iso(datetime.now(timezone.utc))
    updates = {k: v for k, v in {"pv": payload.pv, "earnings": payload.earnings}.items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="Provide pv or earnings")
    await db.user_business.update_one(
        key,
        {"$set": {**key, **updates, "updated_at": now},
         "$setOnInsert": {"created_at": now}},
        upsert=True,
    )
    return {"ok": True}


@api.get("/admin/users/{uid}/business")
async def admin_get_user_business(uid: str, request: Request, season_id: Optional[str] = None):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin", "team_leader"])
    q = {"user_id": uid}
    if season_id:
        q["season_id"] = season_id
    items = await db.user_business.find(q, {"_id": 0}).sort("created_at", -1).to_list(200)
    return items


@api.delete("/admin/users/{uid}/business/{season_id}")
async def admin_delete_user_business(uid: str, season_id: str, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    r = await db.user_business.delete_one({"user_id": uid, "season_id": season_id})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"ok": True}


@api.get("/season/business-totals")
async def season_business_totals(request: Request, season_id: Optional[str] = None):
    """Aggregate PV & Earnings across all members of a season."""
    await get_current_user(request, db)
    if season_id:
        season = await db.seasons.find_one({"season_id": season_id}, {"_id": 0})
    else:
        season = await _current_active_season()
    if not season:
        return {"season": None, "total_pv": 0.0, "total_earnings": 0.0, "member_count": 0}
    agg = await db.user_business.aggregate([
        {"$match": {"season_id": season["season_id"]}},
        {"$group": {"_id": None, "pv": {"$sum": "$pv"}, "earnings": {"$sum": "$earnings"},
                    "members": {"$sum": 1}}},
    ]).to_list(1)
    total_pv_from_ledger = agg[0]["pv"] if agg else 0.0
    total_earnings_from_ledger = agg[0]["earnings"] if agg else 0.0
    # Prefer admin-set season totals when explicitly provided (including 0); fall back to ledger sum.
    season_pv = season.get("total_pv")
    season_earnings = season.get("total_earnings")
    return {
        "season": season,
        "total_pv": round(float(season_pv if season_pv is not None else total_pv_from_ledger), 2),
        "total_earnings": round(float(season_earnings if season_earnings is not None else total_earnings_from_ledger), 2),
        "member_count": agg[0]["members"] if agg else 0,
        "per_member_pv": total_pv_from_ledger,
        "per_member_earnings": total_earnings_from_ledger,
    }


# ---------- Notifications (in-app) ----------
@api.get("/notifications")
async def notifications(request: Request):
    """Return grouped counts + preview items for the notification bell."""
    user = await get_current_user(request, db)
    uid = user["user_id"]
    today_iso = date.today().isoformat()

    # Overdue followups
    overdue = await db.followups.find({
        "user_id": uid, "status": "pending", "due_date": {"$lt": today_iso},
    }, {"_id": 0}).sort("due_date", 1).to_list(20)
    # Due today
    due_today = await db.followups.find({
        "user_id": uid, "status": "pending", "due_date": today_iso,
    }, {"_id": 0}).sort("time_slot", 1).to_list(20)

    # Pending Goals nearing period end
    weekly_start = (date.today() - timedelta(days=date.today().weekday())).isoformat()
    monthly_start = date.today().replace(day=1).isoformat()
    pending_weekly = await db.goals.count_documents({
        "user_id": uid, "status": "active", "period": "weekly", "period_start": weekly_start,
    })
    pending_monthly = await db.goals.count_documents({
        "user_id": uid, "status": "active", "period": "monthly", "period_start": monthly_start,
    })

    # Tasks pending
    tasks_pending = await db.tasks.count_documents({
        "assigned_to": uid, "status": "pending",
    })
    tasks_overdue = await db.tasks.count_documents({
        "assigned_to": uid, "status": "pending", "due_date": {"$lt": today_iso},
    })

    return {
        "unread": len(overdue) + len(due_today) + pending_weekly + pending_monthly + tasks_overdue,
        "followups_overdue": overdue,
        "followups_due_today": due_today,
        "goals_pending_weekly": pending_weekly,
        "goals_pending_monthly": pending_monthly,
        "tasks_pending": tasks_pending,
        "tasks_overdue": tasks_overdue,
    }


# ---------- Season History Archive ----------
@api.post("/admin/seasons/{season_id}/finalize")
async def admin_finalize_season(season_id: str, request: Request):
    """Snapshot the individual + team standings of a season for permanent history."""
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    season = await db.seasons.find_one({"season_id": season_id}, {"_id": 0})
    if not season:
        raise HTTPException(status_code=404, detail="Season not found")
    individual = await spartans_individual(request, season_id=season_id, limit=500)
    team = await spartans_team_league(request, season_id=season_id)
    business = await db.user_business.find({"season_id": season_id}, {"_id": 0}).to_list(5000)
    now = _iso(datetime.now(timezone.utc))
    snapshot = {
        "snapshot_id": str(uuid.uuid4()),
        "season_id": season_id, "season": season,
        "individual_ranking": individual.get("rows", []),
        "team_ranking": team.get("teams", []),
        "business_ledger": business,
        "finalized_at": now,
        "finalized_by": user["user_id"],
    }
    await db.season_snapshots.update_one(
        {"season_id": season_id}, {"$set": snapshot}, upsert=True,
    )
    return {"ok": True, "snapshot_id": snapshot["snapshot_id"]}


@api.get("/season-history")
async def list_season_history(request: Request):
    await get_current_user(request, db)
    snaps = await db.season_snapshots.find({}, {"_id": 0}).sort("finalized_at", -1).to_list(50)
    return snaps


@api.get("/season-history/{season_id}")
async def get_season_snapshot(season_id: str, request: Request):
    await get_current_user(request, db)
    snap = await db.season_snapshots.find_one({"season_id": season_id}, {"_id": 0})
    if not snap:
        raise HTTPException(status_code=404, detail="No snapshot for this season")
    return snap


@api.post("/admin/wipe-demo-data")
async def admin_wipe_demo_data(request: Request):
    """Nuke: prospects, followups, missions, tasks, attendance, xp_events, challenges, rewards, redemptions,
       goals, user_business, all users except super_admin. Preserves teams, weekly_events, seasons."""
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    collections = ["prospects", "followups", "missions", "tasks", "attendance", "event_attendance",
                   "xp_events", "challenges", "rewards", "redemptions", "goals", "user_business",
                   "season_snapshots", "files", "notifications"]
    results = {}
    for c in collections:
        r = await db[c].delete_many({})
        results[c] = r.deleted_count
    # Delete all non-admin users
    r = await db.users.delete_many({"role": {"$ne": "super_admin"}})
    results["users"] = r.deleted_count
    # Reset admin's XP & streaks
    await db.users.update_many({"role": "super_admin"},
                               {"$set": {"xp": 0, "level": 1, "streak_current": 0,
                                         "streak_longest": 0, "last_checkin_date": None,
                                         "badges": [], "position_badges": []}})
    return {"ok": True, "wiped": results}


# ---------- Goals ----------
def _goal_period_start(period: str) -> str:
    today = date.today()
    if period == "weekly":
        return (today - timedelta(days=today.weekday())).isoformat()
    return today.replace(day=1).isoformat()


@api.get("/goals")
async def list_goals(request: Request):
    user = await get_current_user(request, db)
    items = await db.goals.find({"user_id": user["user_id"]}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return items


@api.post("/goals")
async def create_goal(payload: GoalIn, request: Request):
    user = await get_current_user(request, db)
    doc = payload.model_dump()
    doc.update({
        "goal_id": str(uuid.uuid4()),
        "user_id": user["user_id"],
        "progress": 0,
        "status": "active",
        "period_start": _goal_period_start(payload.period),
        "completed_at": None,
        "created_at": _iso(datetime.now(timezone.utc)),
    })
    await db.goals.insert_one(doc)
    return _clean(doc)


@api.patch("/goals/{gid}/progress")
async def update_goal_progress(gid: str, payload: GoalProgress, request: Request):
    user = await get_current_user(request, db)
    g = await db.goals.find_one({"goal_id": gid, "user_id": user["user_id"]}, {"_id": 0})
    if not g:
        raise HTTPException(status_code=404, detail="Goal not found")
    if g.get("status") == "completed":
        raise HTTPException(status_code=400, detail="Already completed")
    new_progress = min(payload.progress, g["target"])
    completed = new_progress >= g["target"]
    updates = {"progress": new_progress}
    xp_awarded = None
    if completed:
        updates["status"] = "completed"
        updates["completed_at"] = _iso(datetime.now(timezone.utc))
        xp_awarded = await award_xp(user["user_id"], int(g.get("xp_reward", 0)), f"goal:{gid}")
    await db.goals.update_one({"goal_id": gid}, {"$set": updates})
    updated = await db.goals.find_one({"goal_id": gid}, {"_id": 0})
    return {"goal": updated, "xp": xp_awarded}


@api.delete("/goals/{gid}")
async def delete_goal(gid: str, request: Request):
    user = await get_current_user(request, db)
    g = await db.goals.find_one({"goal_id": gid, "user_id": user["user_id"]}, {"_id": 0, "assigned_by_admin": 1})
    if g is None:
        raise HTTPException(status_code=404, detail="Goal not found")
    if g.get("assigned_by_admin"):
        raise HTTPException(status_code=403, detail="HQ-assigned goals can only be removed from Goal Settings by an admin")
    await db.goals.delete_one({"goal_id": gid, "user_id": user["user_id"]})
    return {"ok": True}


# ---------- Goal Templates (Super Admin — pushes goals to every member) ----------
class GoalTemplateIn(BaseModel):
    title: str = Field(min_length=1, max_length=120)
    target: int = Field(gt=0)
    xp_reward: int = Field(ge=0, default=0)
    period: Literal["weekly", "monthly"]
    start_date: str  # YYYY-MM-DD
    end_date: str    # YYYY-MM-DD
    active: bool = True

    @field_validator("start_date", "end_date")
    @classmethod
    def v_date(cls, v):
        try:
            date.fromisoformat(v)
        except Exception:
            raise ValueError("Date must be YYYY-MM-DD")
        return v


class GoalTemplateUpdate(BaseModel):
    title: Optional[str] = Field(default=None, min_length=1, max_length=120)
    target: Optional[int] = Field(default=None, gt=0)
    xp_reward: Optional[int] = Field(default=None, ge=0)
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    active: Optional[bool] = None


async def _sync_template_to_users(template: dict):
    """For an active template, ensure every active member has a matching goal doc."""
    if not template.get("active"):
        return 0
    users = await db.users.find({"active": {"$ne": False}}, {"_id": 0, "user_id": 1}).to_list(5000)
    created = 0
    now = _iso(datetime.now(timezone.utc))
    for u in users:
        exists = await db.goals.find_one(
            {"user_id": u["user_id"], "template_id": template["template_id"]},
            {"_id": 0, "goal_id": 1},
        )
        if exists:
            continue
        await db.goals.insert_one({
            "goal_id": str(uuid.uuid4()),
            "template_id": template["template_id"],
            "user_id": u["user_id"],
            "title": template["title"],
            "target": template["target"],
            "xp_reward": template["xp_reward"],
            "period": template["period"],
            "period_start": template["start_date"],
            "period_end": template["end_date"],
            "progress": 0,
            "status": "active",
            "completed_at": None,
            "assigned_by_admin": True,
            "created_at": now,
        })
        created += 1
    return created


@api.get("/admin/goal-templates")
async def admin_list_goal_templates(request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    items = await db.goal_templates.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    # Attach live counts
    for t in items:
        t["assigned_count"] = await db.goals.count_documents({"template_id": t["template_id"]})
        t["completed_count"] = await db.goals.count_documents({
            "template_id": t["template_id"], "status": "completed",
        })
    return items


@api.post("/admin/goal-templates")
async def admin_create_goal_template(payload: GoalTemplateIn, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    if date.fromisoformat(payload.end_date) < date.fromisoformat(payload.start_date):
        raise HTTPException(status_code=400, detail="End date must be on or after start date")
    doc = payload.model_dump()
    doc.update({
        "template_id": str(uuid.uuid4()),
        "created_at": _iso(datetime.now(timezone.utc)),
        "created_by": user["user_id"],
    })
    await db.goal_templates.insert_one(doc)
    assigned = await _sync_template_to_users(doc)
    return {"template": _clean(doc), "assigned_to": assigned}


@api.patch("/admin/goal-templates/{template_id}")
async def admin_update_goal_template(template_id: str, payload: GoalTemplateUpdate, request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    t = await db.goal_templates.find_one({"template_id": template_id}, {"_id": 0})
    if not t:
        raise HTTPException(status_code=404, detail="Template not found")
    updates = {k: v for k, v in payload.model_dump(exclude_unset=True).items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    await db.goal_templates.update_one({"template_id": template_id}, {"$set": updates})
    # Propagate metadata edits to non-completed user goals
    cascade = {k: v for k, v in updates.items()
               if k in ("title", "target", "xp_reward", "start_date", "end_date")}
    if cascade:
        remap = {"start_date": "period_start", "end_date": "period_end"}
        cascade = {remap.get(k, k): v for k, v in cascade.items()}
        await db.goals.update_many(
            {"template_id": template_id, "status": {"$ne": "completed"}},
            {"$set": cascade},
        )
    # If newly activated, backfill for any existing users
    if updates.get("active") is True:
        merged = {**t, **updates}
        await _sync_template_to_users(merged)
    return {"ok": True}


@api.delete("/admin/goal-templates/{template_id}")
async def admin_delete_goal_template(template_id: str, request: Request, cascade: bool = True):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    r = await db.goal_templates.delete_one({"template_id": template_id})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    removed = 0
    if cascade:
        # Only delete open (non-completed) goals so completed history + XP remain intact
        result = await db.goals.delete_many({
            "template_id": template_id, "status": {"$ne": "completed"},
        })
        removed = result.deleted_count
    return {"ok": True, "removed_open_goals": removed}


@api.post("/admin/goal-templates/{template_id}/resync")
async def admin_resync_goal_template(template_id: str, request: Request):
    """Push any active template to newly added members."""
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    t = await db.goal_templates.find_one({"template_id": template_id}, {"_id": 0})
    if not t:
        raise HTTPException(status_code=404, detail="Template not found")
    n = await _sync_template_to_users(t)
    return {"ok": True, "assigned_to": n}


@api.get("/")
async def root():
    return {"message": "Spartans Growth League API", "version": "1.0"}


@api.post("/admin/dedupe-data")
async def admin_dedupe_data(request: Request):
    user = await get_current_user(request, db)
    require_role(user, ["super_admin"])
    breakdown = await dedupe_all_collections(return_breakdown=True)
    total = sum(breakdown.values())
    return {"ok": True, "duplicates_removed": total, "breakdown": breakdown}


app.include_router(api)

# --- CORS (security hardened) ------------------------------------------------
_raw_origins = os.environ.get("CORS_ORIGINS", "").strip()
if not _raw_origins or _raw_origins == "*":
    # Never combine wildcard with credentials — silently degrade to a safe default
    # rather than fail deploy. Emergent preview + localhost dev.
    _default = os.environ.get("EMERGENT_APP_URL") or "http://localhost:3000"
    _allowed_origins = [o.strip() for o in _default.split(",") if o.strip()]
    logger.warning(
        f"CORS_ORIGINS not configured or wildcard — falling back to {_allowed_origins}. "
        "Set CORS_ORIGINS in .env to your production domain(s)."
    )
else:
    _allowed_origins = [o.strip() for o in _raw_origins.split(",") if o.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=_allowed_origins,
    allow_methods=["GET", "POST", "PATCH", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "X-Requested-With"],
)


# ---------- Startup ----------
async def _dedupe_collection(collection, key_fields: list, keep: str = "first"):
    """Remove duplicate documents based on a composite key. Returns count removed."""
    pipeline = [
        {"$group": {"_id": {f: f"${f}" for f in key_fields}, "ids": {"$push": "$_id"}, "n": {"$sum": 1}}},
        {"$match": {"n": {"$gt": 1}}},
    ]
    removed = 0
    async for group in collection.aggregate(pipeline):
        ids = group["ids"]
        # Keep the first, delete the rest (or last if keep=="last")
        to_drop = ids[1:] if keep == "first" else ids[:-1]
        if to_drop:
            r = await collection.delete_many({"_id": {"$in": to_drop}})
            removed += r.deleted_count
    return removed


async def dedupe_all_collections(return_breakdown: bool = False):
    """One-time / repeatable clean of any duplicate rows before unique indexes can be enforced."""
    dedupes = [
        (db.users, ["email"]),
        (db.users, ["user_id"]),
        (db.teams, ["name"]),
        (db.teams, ["team_id"]),
        (db.event_attendance, ["user_id", "event_id", "event_date"]),
        (db.checkins, ["user_id", "date"]),
        (db.goals, ["template_id", "user_id"]),
        (db.rewards, ["reward_id"]),
        (db.seasons, ["season_id"]),
        (db.tasks, ["task_id"]),
        (db.missions, ["mission_id"]),
        (db.goal_templates, ["template_id"]),
    ]
    breakdown: dict = {}
    total = 0
    for coll, keys in dedupes:
        try:
            n = await _dedupe_collection(coll, keys)
            if n:
                logger.info(f"Deduped {coll.name}: removed {n} duplicates on {keys}")
            breakdown[f"{coll.name}:{'+'.join(keys)}"] = n
            total += n
        except Exception as e:
            logger.warning(f"Dedupe skipped for {coll.name} on {keys}: {e}")
            breakdown[f"{coll.name}:{'+'.join(keys)}"] = -1  # sentinel for skipped
    return breakdown if return_breakdown else total


async def seed_admin_and_indexes():
    # Dedupe first so unique index creation doesn't fail on legacy duplicates
    await dedupe_all_collections()
    await db.users.create_index("email", unique=True)
    await db.users.create_index("user_id", unique=True)
    await db.users.create_index([("xp", -1)])
    await db.checkins.create_index([("user_id", 1), ("date", 1)], unique=True)
    await db.prospects.create_index([("user_id", 1), ("status", 1)])
    await db.followups.create_index([("user_id", 1), ("due_date", 1)])
    await db.attendance.create_index([("user_id", 1), ("event_date", -1)])
    await db.missions.create_index("mission_id", unique=True)
    await db.missions.create_index([("user_id", 1), ("created_at", -1)])
    await db.weekly_events.create_index("event_id", unique=True)
    await db.event_attendance.create_index([("user_id", 1), ("event_id", 1), ("event_date", 1)], unique=True)
    await db.event_attendance.create_index([("event_date", 1)])
    await db.seasons.create_index("season_id", unique=True)
    await db.tasks.create_index("task_id", unique=True)
    await db.tasks.create_index([("assigned_to", 1), ("due_date", 1)])
    await db.rewards.create_index("reward_id", unique=True)
    await db.rewards.create_index([("active", 1), ("cost_xp", 1)])
    await db.redemptions.create_index("redemption_id", unique=True)
    await db.redemptions.create_index([("user_id", 1), ("created_at", -1)])
    await db.goals.create_index([("user_id", 1), ("created_at", -1)])
    await db.goals.create_index("goal_id", unique=True)
    await db.challenges.create_index([("start_date", 1), ("end_date", 1)])
    await db.challenge_progress.create_index([("user_id", 1), ("challenge_id", 1)], unique=True)
    await db.xp_events.create_index([("user_id", 1), ("created_at", -1)])
    await db.user_sessions.create_index("session_token", unique=True)
    await db.password_reset_tokens.create_index("expires_at", expireAfterSeconds=0)
    await db.login_attempts.create_index("identifier")
    await db.teams.create_index("name", unique=True)
    await db.teams.create_index("team_id", unique=True)
    await db.users.create_index("team_id")

    async def upsert_seed_user(email, password, name, role, team, team_id=None, extras=None):
        existing = await db.users.find_one({"email": email})
        if existing is None:
            doc = {
                "user_id": f"user_{uuid.uuid4().hex[:12]}", "email": email, "name": name,
                "password_hash": hash_password(password), "role": role,
                "avatar_url": None, "picture": None, "xp": 0, "level": 1,
                "streak_current": 0, "streak_longest": 0, "last_checkin_date": None,
                "team": team, "team_id": team_id, "badges": [],
                "created_at": _iso(datetime.now(timezone.utc)), "active": True,
            }
            if extras:
                doc.update(extras)
            await db.users.insert_one(doc)
        else:
            updates = {"role": role, "name": name, "team": team, "team_id": team_id}
            if not existing.get("password_hash") or not verify_password(password, existing["password_hash"]):
                updates["password_hash"] = hash_password(password)
            if extras:
                # Only fill fields that are currently missing to avoid overwriting user edits.
                for k, v in extras.items():
                    if not existing.get(k):
                        updates[k] = v
            await db.users.update_one({"email": email}, {"$set": updates})

    # Seed teams first (idempotent by name)
    async def upsert_team(name):
        existing = await db.teams.find_one({"name": name})
        if existing:
            return existing["team_id"]
        tid = str(uuid.uuid4())
        await db.teams.insert_one({
            "team_id": tid, "name": name, "leader_id": None,
            "created_at": _iso(datetime.now(timezone.utc)),
        })
        return tid

    team_ids = {
        "Command": await upsert_team("Command"),
        "SPARTANS": await upsert_team("SPARTANS"),
        "Alpha": await upsert_team("Alpha"),
        "Bravo": await upsert_team("Bravo"),
        "Delta": await upsert_team("Delta"),
    }

    # Backfill team_id for users with a team name but no team_id (from earlier seeds)
    for name, tid in team_ids.items():
        await db.users.update_many(
            {"team": name, "$or": [{"team_id": None}, {"team_id": {"$exists": False}}]},
            {"$set": {"team_id": tid}},
        )

    seed_profile = {
        "phone": "9876543210", "bio": "Forged in discipline.",
        "dob": "1985-05-10", "gender": "male", "marital_status": "single",
        "city": "Athens", "state": "Attica", "club_type": "Elite",
        "favourite_food": "Steak", "favourite_place": "Sparta", "favourite_hobby": "Training",
        "joining_date": "2024-01-01", "nexus_id": "BC-ADMIN-001",
    }

    # Only super_admin is seeded. All other members created via signup / add-member flows.
    await upsert_seed_user(os.environ.get("ADMIN_EMAIL", "admin@spartans.com"),
                           os.environ.get("ADMIN_PASSWORD", "Spartan123!"),
                           "Spartan Commander", "super_admin", "Command", team_ids["Command"],
                           extras=seed_profile)

    # Seed default 3 weekly events (Tue/Thu/Sat) if empty
    we_count = await db.weekly_events.count_documents({})
    if we_count == 0:
        await db.weekly_events.insert_many([
            {"event_id": str(uuid.uuid4()), "name": "Believer Season Meeting", "weekday": 0,
             "is_believer": True, "active": True,
             "created_at": _iso(datetime.now(timezone.utc))},
            {"event_id": str(uuid.uuid4()), "name": "MCM (Meta Champion Meet)", "weekday": 3,
             "is_believer": False, "active": True,
             "created_at": _iso(datetime.now(timezone.utc))},
            {"event_id": str(uuid.uuid4()), "name": "Spartans Team Meeting", "weekday": 5,
             "is_believer": False, "active": True,
             "created_at": _iso(datetime.now(timezone.utc))},
        ])


@app.on_event("startup")
async def on_startup():
    try:
        await seed_admin_and_indexes()
        logger.info("Startup: seeded admin, users, challenges & indexes")
    except Exception as e:
        logger.exception("Startup seed error: %s", e)


@app.on_event("shutdown")
async def on_shutdown():
    client.close()
